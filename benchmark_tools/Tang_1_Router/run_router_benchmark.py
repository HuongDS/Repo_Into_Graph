import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
import threading
import time
import requests
import urllib3
import openpyxl
import os

# Tắt cảnh báo InsecureRequestWarning khi gọi https localhost không chứng chỉ
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class RouterBenchmarkApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        self.title("Adaptive Context Router - Benchmark Tool")
        self.geometry("1000x700")
        ctk.set_appearance_mode("light")
        
        self.api_url = "https://localhost:55060/api/test/test-router"
        self.default_file = "Checklist_Kiem_Dinh_Tang_1_Router.xlsx"
        self.result_file = "Checklist_Kiem_Dinh_Tang_1_Router_Results.xlsx"
        
        self.setup_ui()
        
    def setup_ui(self):
        # Định nghĩa Fonts
        self.font_title = ctk.CTkFont(family="Segoe UI", size=26, weight="bold")
        self.font_subtitle = ctk.CTkFont(family="Segoe UI", size=14)
        self.font_text = ctk.CTkFont(family="Segoe UI", size=13)
        self.font_bold = ctk.CTkFont(family="Segoe UI", size=13, weight="bold")
        self.font_dashboard = ctk.CTkFont(family="Segoe UI", size=32, weight="bold")
        
        # --- HEADER & DASHBOARD KHU VỰC ---
        top_frame = ctk.CTkFrame(self, fg_color="transparent")
        top_frame.pack(fill="x", padx=25, pady=(20, 10))
        
        # Bên trái: Tiêu đề
        header_left = ctk.CTkFrame(top_frame, fg_color="transparent")
        header_left.pack(side="left", fill="both", expand=True)
        
        title_label = ctk.CTkLabel(header_left, text="Router Benchmark Engine", font=self.font_title, text_color="#1E3A8A") # Màu xanh navy đậm sang trọng
        title_label.pack(anchor="w")
        
        subtitle = ctk.CTkLabel(header_left, text="Hệ thống đánh giá tự động Tầng 1 - Adaptive Context Router", font=self.font_subtitle, text_color="#64748B")
        subtitle.pack(anchor="w", pady=(2, 0))
        
        # Bên phải: Mini Dashboard (Pass / Fail)
        self.dash_frame = ctk.CTkFrame(top_frame, fg_color="#F8FAFC", corner_radius=12, border_width=1, border_color="#E2E8F0")
        self.dash_frame.pack(side="right", padx=10)
        
        # Pass widget
        pass_box = ctk.CTkFrame(self.dash_frame, fg_color="transparent")
        pass_box.pack(side="left", padx=20, pady=10)
        ctk.CTkLabel(pass_box, text="PASS", font=self.font_bold, text_color="#10B981").pack()
        self.pass_label = ctk.CTkLabel(pass_box, text="0", font=self.font_dashboard, text_color="#059669")
        self.pass_label.pack()
        
        # Fail widget
        fail_box = ctk.CTkFrame(self.dash_frame, fg_color="transparent")
        fail_box.pack(side="left", padx=20, pady=10)
        ctk.CTkLabel(fail_box, text="FAIL", font=self.font_bold, text_color="#EF4444").pack()
        self.fail_label = ctk.CTkLabel(fail_box, text="0", font=self.font_dashboard, text_color="#DC2626")
        self.fail_label.pack()

        # --- CONTROL KHU VỰC ---
        control_frame = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=12, border_width=1, border_color="#E2E8F0")
        control_frame.pack(fill="x", padx=25, pady=10)
        
        ctk.CTkLabel(control_frame, text="File Excel Checklist:", font=self.font_bold, text_color="#334155").pack(side="left", padx=(20, 10), pady=15)
        
        self.file_path_var = ctk.StringVar(value=os.path.abspath(self.default_file) if os.path.exists(self.default_file) else "")
        file_entry = ctk.CTkEntry(control_frame, textvariable=self.file_path_var, font=self.font_text, width=450, fg_color="#F8FAFC", border_color="#CBD5E1")
        file_entry.pack(side="left", padx=10, pady=15)
        
        browse_btn = ctk.CTkButton(control_frame, text="🔍 Browse", font=self.font_bold, command=self.browse_file, width=100, fg_color="#3B82F6", hover_color="#2563EB", corner_radius=8, height=36)
        browse_btn.pack(side="left", padx=10)
        
        self.start_btn = ctk.CTkButton(control_frame, text="▶ CHẠY BENCHMARK", font=ctk.CTkFont(family="Segoe UI", size=14, weight="bold"), command=self.start_benchmark, fg_color="#10B981", hover_color="#059669", corner_radius=8, height=40, width=200)
        self.start_btn.pack(side="right", padx=20)
        
        # --- PROGRESS KHU VỰC ---
        prog_frame = ctk.CTkFrame(self, fg_color="transparent")
        prog_frame.pack(fill="x", padx=25, pady=5)
        
        self.progress_var = ctk.DoubleVar(value=0)
        self.progress_bar = ctk.CTkProgressBar(prog_frame, variable=self.progress_var, height=10, progress_color="#3B82F6", fg_color="#E2E8F0")
        self.progress_bar.pack(fill="x", pady=(0, 5))
        
        self.status_label = ctk.CTkLabel(prog_frame, text="Trạng thái: Sẵn sàng...", font=self.font_text, text_color="#64748B")
        self.status_label.pack(anchor="w")
        
        # --- TABLE KHU VỰC ---
        table_frame = ctk.CTkFrame(self, fg_color="#FFFFFF", corner_radius=12, border_width=1, border_color="#E2E8F0")
        table_frame.pack(fill="both", expand=True, padx=25, pady=(10, 25))
        
        # Tùy chỉnh Style cho Treeview (Sáng, hiện đại)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview.Heading", font=("Segoe UI", 11, "bold"), background="#F1F5F9", foreground="#334155", rowheight=35)
        style.configure("Treeview", font=("Segoe UI", 11), background="#FFFFFF", fieldbackground="#FFFFFF", rowheight=30, borderwidth=0)
        style.map("Treeview", background=[('selected', '#E2E8F0')], foreground=[('selected', '#0F172A')])
        
        columns = ("ma_test", "nhom", "ngon_ngu", "ky_vong", "thuc_te", "status", "latency")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", style="Treeview")
        
        self.tree.heading("ma_test", text="Mã Test")
        self.tree.heading("nhom", text="Nhóm Tiêu Chí")
        self.tree.heading("ngon_ngu", text="Ngôn Ngữ")
        self.tree.heading("ky_vong", text="Kỳ Vọng (V/S/R)")
        self.tree.heading("thuc_te", text="Thực Tế (V/S/R)")
        self.tree.heading("status", text="Kết Quả")
        self.tree.heading("latency", text="Độ Trễ (ms)")
        
        self.tree.column("ma_test", width=100, anchor="center")
        self.tree.column("nhom", width=180)
        self.tree.column("ngon_ngu", width=100, anchor="center")
        self.tree.column("ky_vong", width=200)
        self.tree.column("thuc_te", width=200)
        self.tree.column("status", width=100, anchor="center")
        self.tree.column("latency", width=100, anchor="center")
        
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        
        self.tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        scrollbar.pack(side="right", fill="y", padx=(0, 10), pady=10)
        
        # Tag cấu hình màu nền
        self.tree.tag_configure('PASS', background='#ECFDF5', foreground='#065F46')
        self.tree.tag_configure('FAIL', background='#FEF2F2', foreground='#991B1B')
        
        # Double click event for details
        self.tree.bind("<Double-1>", self.on_tree_double_click)
        
        # Lưu test cases metadata để tra cứu
        self.tc_metadata = {}
        
    def on_tree_double_click(self, event):
        item_id = self.tree.focus()
        if not item_id:
            return
        
        values = self.tree.item(item_id, 'values')
        ma_test = values[0]
        
        if ma_test not in self.tc_metadata:
            return
            
        tc = self.tc_metadata[ma_test]
        status = values[5]
        latency = values[6]
        
        self.show_detail_modal(tc, status, latency)
        
    def show_detail_modal(self, tc, status, latency):
        modal = ctk.CTkToplevel(self)
        modal.title(f"Chi tiết Test Case - {tc['ma_test']}")
        modal.geometry("850x650")
        modal.transient(self)
        modal.grab_set()
        
        # Tiêu đề
        header = ctk.CTkFrame(modal, fg_color="#1E293B", corner_radius=0)
        header.pack(fill="x")
        ctk.CTkLabel(header, text=f"Test Case: {tc['ma_test']}", font=ctk.CTkFont(size=20, weight="bold"), text_color="white").pack(pady=15)
        
        main_frame = ctk.CTkScrollableFrame(modal, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Result overview
        res_frame = ctk.CTkFrame(main_frame, fg_color="#F8FAFC", corner_radius=10, border_width=1, border_color="#E2E8F0")
        res_frame.pack(fill="x", pady=(0, 20))
        
        col1 = ctk.CTkFrame(res_frame, fg_color="transparent")
        col1.pack(side="left", padx=20, pady=15)
        ctk.CTkLabel(col1, text="Kết quả:", font=self.font_bold).pack(anchor="w")
        color = "#10B981" if status == "PASS" else "#EF4444"
        ctk.CTkLabel(col1, text=status, font=ctk.CTkFont(size=18, weight="bold"), text_color=color).pack(anchor="w")
        
        col2 = ctk.CTkFrame(res_frame, fg_color="transparent")
        col2.pack(side="left", padx=20, pady=15)
        ctk.CTkLabel(col2, text="Độ trễ:", font=self.font_bold).pack(anchor="w")
        ctk.CTkLabel(col2, text=f"{latency} ms", font=ctk.CTkFont(size=16)).pack(anchor="w")
        
        col3 = ctk.CTkFrame(res_frame, fg_color="transparent")
        col3.pack(side="left", padx=20, pady=15)
        ctk.CTkLabel(col3, text="Nhóm:", font=self.font_bold).pack(anchor="w")
        ctk.CTkLabel(col3, text=tc['nhom'], font=ctk.CTkFont(size=14)).pack(anchor="w")
        
        # Code Preview
        ctk.CTkLabel(main_frame, text="Mã nguồn (Source Code):", font=self.font_bold).pack(anchor="w")
        textbox = ctk.CTkTextbox(main_frame, height=250, font=("Consolas", 13), fg_color="#F1F5F9", text_color="#0F172A", border_color="#CBD5E1", border_width=1)
        textbox.pack(fill="x", pady=5)
        textbox.insert("0.0", tc.get('code', 'Không có mã nguồn'))
        textbox.configure(state="disabled")
        
        # Details grid
        grid = ctk.CTkFrame(main_frame, fg_color="transparent")
        grid.pack(fill="x", pady=15)
        
        def add_row(parent, label_text, kv, tt):
            row = ctk.CTkFrame(parent, fg_color="transparent")
            row.pack(fill="x", pady=2)
            ctk.CTkLabel(row, text=label_text, width=150, anchor="w", font=self.font_bold).pack(side="left")
            ctk.CTkLabel(row, text=f"Kỳ vọng: {kv}", width=250, anchor="w", text_color="#475569").pack(side="left")
            ctk.CTkLabel(row, text=f"Thực tế: {tt}", width=250, anchor="w", text_color="#0F172A", font=self.font_bold).pack(side="left")
            
        add_row(grid, "Logic LSLOC:", tc['kv_sloc'], tc.get('tt_sloc', ''))
        add_row(grid, "Cyclomatic V(G):", tc['kv_vg'], tc.get('tt_vg', ''))
        add_row(grid, "Quyết định (Route):", tc['kv_route'], tc.get('tt_route', ''))
        
        if tc.get('ghi_chu'):
            ctk.CTkLabel(main_frame, text="Ghi chú (Lỗi):", font=self.font_bold, text_color="#DC2626").pack(anchor="w", pady=(10, 0))
            ctk.CTkLabel(main_frame, text=tc['ghi_chu'], wraplength=700, justify="left", text_color="#DC2626").pack(anchor="w")
        
    def browse_file(self):
        filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            self.file_path_var.set(filepath)
            
    def start_benchmark(self):
        file_path = self.file_path_var.get()
        if not file_path or not os.path.exists(file_path):
            messagebox.showerror("Lỗi", "Không tìm thấy file checklist.")
            return
            
        self.start_btn.configure(state="disabled")
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # Chạy thread để không block UI
        thread = threading.Thread(target=self.run_benchmark_process, args=(file_path,))
        thread.start()
        
    def run_benchmark_process(self, file_path):
        try:
            self.pass_label.configure(text="0")
            self.fail_label.configure(text="0")
            self.status_label.configure(text="Trạng thái: Đang đọc file Excel...")
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Đọc từ dòng 6 trở đi
            test_cases = []
            for row_idx in range(6, ws.max_row + 1):
                ma_test = ws.cell(row=row_idx, column=1).value
                # Bỏ qua các dòng tiêu đề (không có mã hoặc không bắt đầu bằng TC_)
                if not ma_test or not str(ma_test).strip().startswith('TC_'):
                    continue
                    
                nhom = ws.cell(row=row_idx, column=2).value
                ngon_ngu = ws.cell(row=row_idx, column=4).value
                code = ws.cell(row=row_idx, column=5).value
                kv_vg = ws.cell(row=row_idx, column=6).value
                kv_sloc = ws.cell(row=row_idx, column=7).value
                kv_route = ws.cell(row=row_idx, column=8).value
                
                test_cases.append({
                    "row": row_idx,
                    "ma_test": str(ma_test).strip(),
                    "nhom": str(nhom) if nhom else "",
                    "ngon_ngu": str(ngon_ngu).strip() if ngon_ngu else "",
                    "code": str(code) if code else "",
                    "kv_vg": str(kv_vg).strip() if kv_vg else "",
                    "kv_sloc": str(kv_sloc).strip() if kv_sloc else "",
                    "kv_route": str(kv_route).strip() if kv_route else ""
                })
                
            total = len(test_cases)
            pass_count = 0
            fail_count = 0
            
            for i, tc in enumerate(test_cases):
                self.status_label.configure(text=f"Trạng thái: Đang chạy {tc['ma_test']} ({i+1}/{total})...")
                
                # Default values for actual
                tt_vg = ""
                tt_sloc = ""
                tt_route = ""
                status = "FAIL"
                ghi_chu = ""
                latency_ms = 0
                
                start_time = time.time()
                try:
                    lang = tc['ngon_ngu'].split('/')[0].strip()
                    payload = {
                        "sourceCode": tc['code'],
                        "language": lang
                    }
                    response = requests.post(self.api_url, json=payload, verify=False, timeout=10)
                    latency_ms = int((time.time() - start_time) * 1000)
                    
                    if response.status_code == 200:
                        data = response.json()
                        
                        if not data.get("isValidSyntax", False):
                            msg = data.get("message", "")
                            if "Syntax Error" in msg or "cú pháp" in msg.lower():
                                tt_vg = "has_error=True"
                                tt_sloc = "has_error=True"
                                tt_route = "ROUTE_RAW_CODE (Fallback)"
                            elif "Bad Request" in msg or "400" in msg:
                                tt_vg = "N/A"
                                tt_sloc = "N/A"
                                tt_route = "HTTP 400 Bad Request"
                            else:
                                tt_vg = "N/A"
                                tt_sloc = "N/A"
                                tt_route = "Exception"
                            ghi_chu = msg
                        else:
                            tt_vg = str(data.get("vg", ""))
                            tt_sloc = str(data.get("sloc", ""))
                            # Map int enum to string
                            route_enum = data.get("selectedRoute", 0)
                            tt_route = "ROUTE_HYBRID" if route_enum == 1 else "ROUTE_RAW_CODE"
                            ghi_chu = data.get("message", "")
                            
                    elif response.status_code == 400:
                        tt_route = "HTTP 400 Bad Request"
                        ghi_chu = "Bad Request / Unsupported"
                        tt_vg = "N/A"
                        tt_sloc = "N/A"
                    else:
                        tt_route = f"HTTP {response.status_code}"
                        ghi_chu = "Lỗi Server"
                        tt_vg = "N/A"
                        tt_sloc = "N/A"
                        
                except Exception as e:
                    latency_ms = int((time.time() - start_time) * 1000)
                    tt_route = "Exception"
                    ghi_chu = str(e)
                    tt_vg = "N/A"
                    tt_sloc = "N/A"
                
                # Cập nhật kết quả thực tế vào metadata để hiển thị Modal
                tc['tt_vg'] = tt_vg
                tc['tt_sloc'] = tt_sloc
                tc['tt_route'] = tt_route
                tc['ghi_chu'] = ghi_chu
                self.tc_metadata[tc['ma_test']] = tc
                
                # Logic Assert
                is_pass = True
                
                # Check V(G)
                if tc['kv_vg'] != 'N/A' and tc['kv_vg'] != tt_vg:
                    is_pass = False
                
                # Check SLOC
                if tc['kv_sloc'] != 'N/A' and tc['kv_sloc'] != tt_sloc:
                    is_pass = False
                    
                # Check Route
                if tc['kv_route'] != 'N/A' and tc['kv_route'] not in tt_route and tt_route not in tc['kv_route']:
                    is_pass = False
                    
                # Latency rule TC_SYS_03
                if tc['ma_test'] == 'TC_SYS_03':
                    tt_vg = "N/A"
                    tt_sloc = "N/A"
                    if latency_ms > 15:
                        is_pass = False
                        ghi_chu = f"Latency {latency_ms}ms > 15ms"
                        tt_route = f"Latency {latency_ms}ms > 15ms"
                    else:
                        is_pass = True
                        tt_route = f"Latency < 15ms"
                        
                if is_pass:
                    status = "PASS"
                    pass_count += 1
                    self.pass_label.configure(text=str(pass_count))
                else:
                    status = "FAIL"
                    fail_count += 1
                    self.fail_label.configure(text=str(fail_count))
                    
                # Ghi vào file excel
                from openpyxl.cell.cell import MergedCell
                def safe_write(r, c, val):
                    cell = ws.cell(row=r, column=c)
                    if not isinstance(cell, MergedCell):
                        cell.value = val

                safe_write(tc['row'], 9, tt_vg)
                safe_write(tc['row'], 10, tt_sloc)
                safe_write(tc['row'], 11, tt_route)
                safe_write(tc['row'], 12, status)
                safe_write(tc['row'], 13, ghi_chu)
                
                # Update UI table
                kv_text = f"V:{tc['kv_vg']} | S:{tc['kv_sloc']} | {tc['kv_route']}"
                tt_text = f"V:{tt_vg} | S:{tt_sloc} | {tt_route}"
                
                self.tree.insert("", "end", values=(
                    tc['ma_test'], tc['nhom'], tc['ngon_ngu'],
                    kv_text, tt_text, status, str(latency_ms)
                ), tags=(status,))
                
                self.progress_var.set((i + 1) / total)
                
            # Lưu file kết quả
            self.status_label.configure(text="Trạng thái: Đang lưu file kết quả...")
            
            try:
                wb.save(self.result_file)
                self.status_label.configure(text=f"Trạng thái: Hoàn thành! Đã lưu kết quả ra file {self.result_file}", text_color="#10B981")
            except Exception as e:
                self.status_label.configure(text=f"Lỗi khi lưu file: {str(e)}", text_color="#EF4444")
            
        except Exception as e:
            self.status_label.configure(text=f"Lỗi hệ thống: {str(e)}", text_color="#EF4444")
            messagebox.showerror("Lỗi", str(e))
        finally:
            self.start_btn.configure(state="normal")

if __name__ == "__main__":
    app = RouterBenchmarkApp()
    app.mainloop()
