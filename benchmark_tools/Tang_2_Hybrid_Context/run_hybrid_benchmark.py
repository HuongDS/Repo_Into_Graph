import customtkinter as ctk
from tkinter import filedialog, messagebox, ttk
import threading
import time
import re
import requests
import urllib3
import openpyxl
import os
import json

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ─────────────────────────────────────────────
# HELPER: Parse chuỗi "SLOC=6, V(G)=3 -> ROUTE_HYBRID"
# ─────────────────────────────────────────────
def parse_tier1_input(s: str) -> dict:
    result = {"sloc": 0, "vg": 0, "routing_decision": "ROUTE_HYBRID"}
    if not s:
        return result
    m_sloc = re.search(r"SLOC\s*=\s*(\d+)", s, re.IGNORECASE)
    m_vg   = re.search(r"V\(G\)\s*=\s*(\d+)", s, re.IGNORECASE)
    m_route = re.search(r"(ROUTE_HYBRID|ROUTE_RAW_CODE)", s, re.IGNORECASE)
    if m_sloc:
        result["sloc"] = int(m_sloc.group(1))
    if m_vg:
        result["vg"] = int(m_vg.group(1))
    if m_route:
        result["routing_decision"] = m_route.group(1).upper()
    return result


class HybridBenchmarkApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Hybrid Context Generator – Benchmark Tool (Tầng 2)")
        self.geometry("1100x740")
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.api_url      = "https://localhost:55060/api/test/test-hybrid-context"
        self.default_file = "Checklist_Test_Tang_2_Handover.xlsx"
        self.result_file  = "Checklist_Test_Tang_2_Handover_Results.xlsx"

        self.tc_metadata: dict = {}
        self._setup_fonts()
        self._setup_ui()

    # ─── FONTS ───────────────────────────────
    def _setup_fonts(self):
        self.f_title     = ctk.CTkFont("Segoe UI", 26, "bold")
        self.f_subtitle  = ctk.CTkFont("Segoe UI", 14)
        self.f_bold      = ctk.CTkFont("Segoe UI", 13, "bold")
        self.f_text      = ctk.CTkFont("Segoe UI", 13)
        self.f_dashboard = ctk.CTkFont("Segoe UI", 34, "bold")

    # ─── UI ──────────────────────────────────
    def _setup_ui(self):
        # ── Header ──────────────────────────
        top = ctk.CTkFrame(self, fg_color="transparent")
        top.pack(fill="x", padx=25, pady=(20, 8))

        left = ctk.CTkFrame(top, fg_color="transparent")
        left.pack(side="left", fill="both", expand=True)
        ctk.CTkLabel(left, text="Hybrid Context Benchmark", font=self.f_title,
                     text_color="#4F46E5").pack(anchor="w")
        ctk.CTkLabel(left, text="Hệ thống kiểm định bàn giao – Tầng 2: Hybrid Context Generator",
                     font=self.f_subtitle, text_color="#64748B").pack(anchor="w", pady=(2, 0))

        # Mini dashboard PASS / FAIL
        dash = ctk.CTkFrame(top, fg_color="#F8FAFC", corner_radius=14,
                            border_width=1, border_color="#E2E8F0")
        dash.pack(side="right", padx=10)

        for label, color, attr in [("PASS", "#10B981", "pass_label"),
                                    ("FAIL", "#EF4444", "fail_label")]:
            box = ctk.CTkFrame(dash, fg_color="transparent")
            box.pack(side="left", padx=22, pady=12)
            ctk.CTkLabel(box, text=label, font=self.f_bold, text_color=color).pack()
            lbl = ctk.CTkLabel(box, text="0", font=self.f_dashboard,
                               text_color=color.replace("B981", "6D5A").replace("4444", "C2626"))
            lbl.pack()
            setattr(self, attr, lbl)

        # ── API URL bar ─────────────────────
        url_frame = ctk.CTkFrame(self, fg_color="#EEF2FF", corner_radius=10,
                                 border_width=1, border_color="#C7D2FE")
        url_frame.pack(fill="x", padx=25, pady=(0, 6))
        ctk.CTkLabel(url_frame, text="🌐 API Endpoint:", font=self.f_bold,
                     text_color="#4338CA").pack(side="left", padx=(16, 8), pady=10)
        self.url_var = ctk.StringVar(value=self.api_url)
        ctk.CTkEntry(url_frame, textvariable=self.url_var, font=self.f_text,
                     width=520, fg_color="white", border_color="#A5B4FC").pack(side="left", pady=10)

        # ── Control bar ─────────────────────
        ctrl = ctk.CTkFrame(self, fg_color="white", corner_radius=12,
                            border_width=1, border_color="#E2E8F0")
        ctrl.pack(fill="x", padx=25, pady=6)

        ctk.CTkLabel(ctrl, text="📄 Checklist Excel:", font=self.f_bold,
                     text_color="#334155").pack(side="left", padx=(18, 8), pady=14)

        default_path = os.path.abspath(self.default_file) if os.path.exists(self.default_file) else ""
        self.file_var = ctk.StringVar(value=default_path)
        ctk.CTkEntry(ctrl, textvariable=self.file_var, font=self.f_text,
                     width=420, fg_color="#F8FAFC", border_color="#CBD5E1").pack(side="left", padx=8, pady=14)

        ctk.CTkButton(ctrl, text="🔍 Browse", font=self.f_bold, command=self._browse,
                      width=100, fg_color="#6366F1", hover_color="#4F46E5",
                      corner_radius=8, height=36).pack(side="left", padx=8)

        self.run_btn = ctk.CTkButton(ctrl, text="▶  CHẠY BENCHMARK",
                                     font=ctk.CTkFont("Segoe UI", 14, "bold"),
                                     command=self._start, fg_color="#10B981",
                                     hover_color="#059669", corner_radius=8,
                                     height=40, width=210)
        self.run_btn.pack(side="right", padx=18)

        # ── Progress ────────────────────────
        prog_frame = ctk.CTkFrame(self, fg_color="transparent")
        prog_frame.pack(fill="x", padx=25, pady=(0, 4))
        self.progress = ctk.CTkProgressBar(prog_frame, height=8,
                                           fg_color="#E2E8F0", progress_color="#6366F1")
        self.progress.set(0)
        self.progress.pack(fill="x")
        self.status_lbl = ctk.CTkLabel(prog_frame, text="Trạng thái: Sẵn sàng",
                                       font=self.f_text, text_color="#64748B")
        self.status_lbl.pack(anchor="w", pady=(3, 0))

        # ── Results table ────────────────────
        tbl_card = ctk.CTkFrame(self, fg_color="white", corner_radius=12,
                                border_width=1, border_color="#E2E8F0")
        tbl_card.pack(fill="both", expand=True, padx=25, pady=(0, 20))

        cols = ("ma_test", "muc_tieu", "ngon_ngu", "ky_vong_cfg",
                "ky_vong_snippet", "ket_qua", "latency_ms")
        col_labels = ("Mã Test", "Mục Tiêu", "Ngôn Ngữ",
                      "Kỳ Vọng CFG (Skeleton)", "Kỳ Vọng Snippets",
                      "Kết Quả", "Latency (ms)")
        col_widths  = (90, 200, 80, 200, 200, 70, 80)

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Treeview",
                        background="white", foreground="#1E293B",
                        rowheight=36, fieldbackground="white",
                        font=("Segoe UI", 12))
        style.configure("Custom.Treeview.Heading",
                        background="#F1F5F9", foreground="#374151",
                        font=("Segoe UI", 12, "bold"), relief="flat")
        style.map("Custom.Treeview",
                  background=[("selected", "#EEF2FF")],
                  foreground=[("selected", "#4338CA")])

        scroll = ttk.Scrollbar(tbl_card, orient="vertical")
        self.tree = ttk.Treeview(tbl_card, columns=cols, show="headings",
                                 yscrollcommand=scroll.set, style="Custom.Treeview")
        scroll.configure(command=self.tree.yview)

        for col, lbl, w in zip(cols, col_labels, col_widths):
            self.tree.heading(col, text=lbl)
            self.tree.column(col, width=w, anchor="center")

        self.tree.pack(side="left", fill="both", expand=True, padx=(10, 0), pady=10)
        scroll.pack(side="right", fill="y", padx=(0, 10), pady=10)

        self.tree.tag_configure("PASS", background="#ECFDF5", foreground="#065F46")
        self.tree.tag_configure("FAIL", background="#FEF2F2", foreground="#991B1B")
        self.tree.bind("<Double-1>", self._on_double_click)

    # ─── EVENTS ──────────────────────────────
    def _browse(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.file_var.set(path)

    def _start(self):
        path = self.file_var.get()
        if not path or not os.path.exists(path):
            messagebox.showerror("Lỗi", "Không tìm thấy file checklist.")
            return
        self.api_url = self.url_var.get().strip()
        self.run_btn.configure(state="disabled")
        for item in self.tree.get_children():
            self.tree.delete(item)
        threading.Thread(target=self._run, args=(path,), daemon=True).start()

    def _on_double_click(self, _event):
        item = self.tree.focus()
        if not item:
            return
        ma = self.tree.item(item, "values")[0]
        if ma in self.tc_metadata:
            self._show_modal(self.tc_metadata[ma])

    # ─── DETAIL MODAL ────────────────────────
    def _show_modal(self, tc: dict):
        status  = tc.get("status", "")
        latency = tc.get("latency_ms", 0)

        modal = ctk.CTkToplevel(self)
        modal.title(f"Chi tiết – {tc['ma_test']}")
        modal.geometry("900x700")
        modal.transient(self)
        modal.grab_set()

        # Header
        hdr = ctk.CTkFrame(modal, fg_color="#312E81", corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text=f"Test Case: {tc['ma_test']}  ·  {tc['muc_tieu']}",
                     font=ctk.CTkFont("Segoe UI", 18, "bold"),
                     text_color="white").pack(pady=14)

        scroll_area = ctk.CTkScrollableFrame(modal, fg_color="#F8FAFC")
        scroll_area.pack(fill="both", expand=True, padx=0, pady=0)

        # Result bar
        res_bar = ctk.CTkFrame(scroll_area, fg_color="white", corner_radius=10,
                               border_width=1, border_color="#E2E8F0")
        res_bar.pack(fill="x", padx=20, pady=(20, 0))
        for label, val, color in [
            ("Kết Quả", status, "#10B981" if status == "PASS" else "#EF4444"),
            ("Độ Trễ",  f"{latency} ms", "#6366F1"),
            ("Ngôn Ngữ", tc["ngon_ngu"], "#0EA5E9"),
        ]:
            col = ctk.CTkFrame(res_bar, fg_color="transparent")
            col.pack(side="left", padx=24, pady=14)
            ctk.CTkLabel(col, text=label, font=self.f_bold, text_color="#64748B").pack(anchor="w")
            ctk.CTkLabel(col, text=val, font=ctk.CTkFont("Segoe UI", 18, "bold"),
                         text_color=color).pack(anchor="w")

        def section(title):
            ctk.CTkLabel(scroll_area, text=title, font=self.f_bold,
                         text_color="#374151").pack(anchor="w", padx=20, pady=(16, 4))

        def code_box(text, height=140):
            tb = ctk.CTkTextbox(scroll_area, height=height, font=("Consolas", 12),
                                fg_color="#F1F5F9", text_color="#1E293B",
                                border_color="#CBD5E1", border_width=1)
            tb.pack(fill="x", padx=20, pady=(0, 4))
            tb.insert("0.0", text or "(trống)")
            tb.configure(state="disabled")
            return tb

        section("📝 Source Code Input")
        code_box(tc.get("code", ""), 150)

        section("📐 Kỳ Vọng CFG Skeleton")
        code_box(tc.get("kv_cfg", ""), 120)

        section("✂️  Kỳ Vọng Critical Snippets")
        code_box(tc.get("kv_snippet", ""), 100)

        section("📤 Payload Đã Gửi Lên API")
        code_box(json.dumps(tc.get("payload_sent", {}), indent=2, ensure_ascii=False), 160)

        section("📥 Response Nhận Về")
        color_resp = "#ECFDF5" if status == "PASS" else "#FEF2F2"
        resp_tb = ctk.CTkTextbox(scroll_area, height=140, font=("Consolas", 12),
                                 fg_color=color_resp, text_color="#1E293B",
                                 border_color="#CBD5E1", border_width=1)
        resp_tb.pack(fill="x", padx=20, pady=(0, 4))
        resp_tb.insert("0.0", json.dumps(tc.get("response_data", {}), indent=2, ensure_ascii=False))
        resp_tb.configure(state="disabled")

        if tc.get("ghi_chu"):
            section("⚠️  Ghi Chú / Lỗi")
            ctk.CTkLabel(scroll_area, text=tc["ghi_chu"], wraplength=820,
                         justify="left", text_color="#DC2626",
                         font=self.f_text).pack(anchor="w", padx=20, pady=(0, 16))

    # ─── BENCHMARK PROCESS ───────────────────
    def _run(self, file_path: str):
        try:
            self.pass_label.configure(text="0")
            self.fail_label.configure(text="0")
            self.status_lbl.configure(text="Trạng thái: Đang đọc file Excel...")

            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # --- Đọc test cases (từ row 5, bỏ header rows 1-4) ---
            test_cases = []
            for row in range(5, ws.max_row + 1):
                ma = ws.cell(row, 1).value
                if not ma or not str(ma).strip().startswith("TC_"):
                    continue
                test_cases.append({
                    "row":        row,
                    "ma_test":    str(ma).strip(),
                    "muc_tieu":   str(ws.cell(row, 2).value or ""),
                    "ngon_ngu":   str(ws.cell(row, 3).value or "java").strip(),
                    "code":       str(ws.cell(row, 4).value or ""),
                    "tier1_str":  str(ws.cell(row, 5).value or ""),
                    "kv_cfg":     str(ws.cell(row, 6).value or ""),
                    "kv_snippet": str(ws.cell(row, 7).value or ""),
                })

            total = len(test_cases)
            pass_count = fail_count = 0

            for i, tc in enumerate(test_cases):
                self.status_lbl.configure(
                    text=f"Trạng thái: Đang chạy {tc['ma_test']} ({i+1}/{total})...")

                # Parse input từ Tầng 1
                t1 = parse_tier1_input(tc["tier1_str"])

                ghi_chu      = ""
                latency_ms   = 0
                response_data: dict = {}
                is_pass      = False

                payload = {
                    "moduleId":        tc["ma_test"],
                    "language":        tc["ngon_ngu"].lower(),
                    "routingDecision": t1["routing_decision"],
                    "metrics": {
                        "sloc":                t1["sloc"],
                        "cyclomaticComplexity": t1["vg"],
                    },
                    "rawSourceCode": tc["code"],
                    "astPayload": {
                        "parserType":   "tree-sitter",
                        "rootNodeType": "",
                        "hasError":     False,
                    },
                }

                t_start = time.time()
                try:
                    resp = requests.post(self.api_url, json=payload,
                                         verify=False, timeout=15)
                    latency_ms = int((time.time() - t_start) * 1000)

                    if resp.status_code == 200:
                        response_data = resp.json()
                        # ── ASSERT ──────────────────────────────────────
                        # Cột kỳ vọng 6: CFG skeleton (so sánh fuzzy)
                        # Cột kỳ vọng 7: snippets (so sánh fuzzy)
                        # Hiện Stub trả về status/message → assert thông tin
                        # cốt lõi có trong message hay không
                        status_val = response_data.get("status", "")
                        msg_val    = response_data.get("message", "")
                        mid_val    = response_data.get("moduleId", "")

                        # Assert 1: moduleId phải được phản chiếu lại
                        if mid_val != tc["ma_test"]:
                            ghi_chu += f"moduleId sai: nhận '{mid_val}', kỳ vọng '{tc['ma_test']}'. "
                            is_pass = False
                        else:
                            is_pass = True

                        # Assert 2: CFG skeleton kỳ vọng — khi Tang 2 implement
                        # sẽ so sánh nội dung trường cfgSkeleton. Hiện stub → skip
                        if tc["kv_cfg"].strip() and "cfgSkeleton" in response_data:
                            if tc["kv_cfg"].strip() not in str(response_data["cfgSkeleton"]):
                                is_pass = False
                                ghi_chu += "CFG skeleton không khớp kỳ vọng. "

                        # Assert 3: Critical snippets — tương tự
                        if tc["kv_snippet"].strip() and "criticalSnippets" in response_data:
                            for line in tc["kv_snippet"].split("\n"):
                                line = line.strip()
                                if line and line not in str(response_data["criticalSnippets"]):
                                    is_pass = False
                                    ghi_chu += f"Snippet thiếu: '{line[:40]}'. "
                                    break

                    elif resp.status_code == 400:
                        is_pass = False
                        ghi_chu = f"HTTP 400: {resp.text[:200]}"
                        response_data = {"error": ghi_chu}
                    else:
                        is_pass = False
                        ghi_chu = f"HTTP {resp.status_code}"
                        response_data = {"error": ghi_chu}

                except Exception as ex:
                    latency_ms = int((time.time() - t_start) * 1000)
                    is_pass = False
                    ghi_chu = str(ex)
                    response_data = {"error": ghi_chu}

                status = "PASS" if is_pass else "FAIL"
                if is_pass:
                    pass_count += 1
                    self.pass_label.configure(text=str(pass_count))
                else:
                    fail_count += 1
                    self.fail_label.configure(text=str(fail_count))

                # Lưu metadata cho modal
                tc.update({
                    "status":        status,
                    "latency_ms":    latency_ms,
                    "payload_sent":  payload,
                    "response_data": response_data,
                    "ghi_chu":       ghi_chu,
                })
                self.tc_metadata[tc["ma_test"]] = tc

                # Ghi vào Excel
                from openpyxl.cell.cell import MergedCell

                def safe_write(r, c, v):
                    cell = ws.cell(row=r, column=c)
                    if not isinstance(cell, MergedCell):
                        cell.value = v

                safe_write(tc["row"], 8, status)

                # Cập nhật bảng kết quả
                cfg_short = (tc["kv_cfg"][:60] + "…") if len(tc["kv_cfg"]) > 60 else tc["kv_cfg"]
                snp_short = (tc["kv_snippet"][:60] + "…") if len(tc["kv_snippet"]) > 60 else tc["kv_snippet"]
                self.tree.insert("", "end", values=(
                    tc["ma_test"], tc["muc_tieu"][:40],
                    tc["ngon_ngu"], cfg_short, snp_short,
                    status, str(latency_ms),
                ), tags=(status,))

                self.progress.set((i + 1) / total)

            # Lưu kết quả
            self.status_lbl.configure(text="Trạng thái: Đang lưu file kết quả...")
            try:
                wb.save(self.result_file)
                self.status_lbl.configure(
                    text=f"✅ Hoàn thành! Đã lưu kết quả ra file '{self.result_file}'",
                    text_color="#10B981")
            except Exception as ex:
                self.status_lbl.configure(
                    text=f"⚠️ Không thể lưu file: {ex}", text_color="#F59E0B")

        except Exception as ex:
            self.status_lbl.configure(text=f"❌ Lỗi hệ thống: {ex}", text_color="#EF4444")
            messagebox.showerror("Lỗi", str(ex))
        finally:
            self.run_btn.configure(state="normal")


if __name__ == "__main__":
    app = HybridBenchmarkApp()
    app.mainloop()
