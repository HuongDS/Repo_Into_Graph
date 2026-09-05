import argparse
import requests
import json
import os
import openpyxl
import time
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

def find_excel_file():
    excel_path = "CFG_vs_Traditional_Benchmark_Template.xlsx"
    if not os.path.exists(excel_path):
        current = os.path.abspath(os.path.curdir)
        while current and current != os.path.dirname(current):
            candidate = os.path.join(current, "CFG_vs_Traditional_Benchmark_Template.xlsx")
            if os.path.exists(candidate):
                return candidate
            current = os.path.dirname(current)
    return excel_path

def append_summary(ws, data):
    start_row = -1
    for r in range(4, 25):
        val = ws.cell(row=r, column=3).value # Business Name
        if val is None or str(val).strip() == "":
            start_row = r
            break
            
    if start_row == -1 or start_row > 23:
        start_row = ws.max_row + 1

    run_num = int((start_row - 2)/2)
    if run_num < 1: run_num = 1

    # Traditional Row
    r_trad = start_row
    ws.cell(row=r_trad, column=1, value=run_num)
    ws.cell(row=r_trad, column=2, value=f"{data['runId']}_TRAD")
    ws.cell(row=r_trad, column=3, value=data['businessName'])
    ws.cell(row=r_trad, column=4, value="Traditional")
    ws.cell(row=r_trad, column=5, value=data['tradTime'])
    ws.cell(row=r_trad, column=6, value=data['tradInputTokens'])
    ws.cell(row=r_trad, column=7, value=data['tradOutputTokens'])
    ws.cell(row=r_trad, column=8, value=data['tradCoverage'])
    ws.cell(row=r_trad, column=9, value=data['tradActiveNodes'])
    ws.cell(row=r_trad, column=10, value=data['tradAccuracy'])
    ws.cell(row=r_trad, column=11, value=data['tradComplexity'])
    ws.cell(row=r_trad, column=12, value=data['difficulty'])
    ws.cell(row=r_trad, column=13, value="Tự động ghi qua CLI")

    # CFG Row
    r_cfg = start_row + 1
    ws.cell(row=r_cfg, column=1, value=run_num)
    ws.cell(row=r_cfg, column=2, value=f"{data['runId']}_CFG")
    ws.cell(row=r_cfg, column=3, value=data['businessName'])
    ws.cell(row=r_cfg, column=4, value="Graph-based (CFG)")
    ws.cell(row=r_cfg, column=5, value=data['cfgTime'])
    ws.cell(row=r_cfg, column=6, value=data['cfgInputTokens'])
    ws.cell(row=r_cfg, column=7, value=data['cfgOutputTokens'])
    ws.cell(row=r_cfg, column=8, value=data['cfgCoverage'])
    ws.cell(row=r_cfg, column=9, value=data['cfgActiveNodes'])
    ws.cell(row=r_cfg, column=10, value=data['cfgAccuracy'])
    ws.cell(row=r_cfg, column=11, value=data['cfgComplexity'])
    ws.cell(row=r_cfg, column=12, value=data['difficulty'])
    ws.cell(row=r_cfg, column=13, value="Tự động ghi qua CLI")

def append_details(ws, run_id, mode, details):
    start_row = -1
    for r in range(4, max(5, ws.max_row + 5)):
        val = ws.cell(row=r, column=1).value
        if val is None or str(val).strip() == "":
            start_row = r
            break
    
    if start_row == -1:
        start_row = ws.max_row + 1

    current_row = start_row
    stt = 1
    for q in details:
        ws.cell(row=current_row, column=1, value=f"{run_id}_{'TRAD' if 'Trad' in mode else 'CFG'}")
        ws.cell(row=current_row, column=2, value=mode)
        ws.cell(row=current_row, column=3, value=stt)
        ws.cell(row=current_row, column=4, value=q.get("question", ""))
        ws.cell(row=current_row, column=5, value=q.get("activeNodes", 0))
        ws.cell(row=current_row, column=6, value=q.get("coverage", 0))
        ws.cell(row=current_row, column=7, value="Đúng" if q.get("isAccurate", False) else "Sai")
        ws.cell(row=current_row, column=8, value=f"{q.get('gateways', 0)} Gateways")
        ws.cell(row=current_row, column=9, value=q.get("evaluationNotes", ""))
        
        current_row += 1
        stt += 1

def run_pipeline(api_url, business_id, num_questions, difficulty, mode):
    print(f"\n--- Bắt đầu chạy Pipeline {mode} ---")
    
    generate_endpoint = f"{api_url}/api/QuestionGenerator/generate-traditional" if mode == "Traditional" else f"{api_url}/api/QuestionGenerator/generate-graph"
    
    # 1. Sinh câu hỏi
    print(f"1. Gọi Sinh câu hỏi ({mode})...")
    start_time = time.time()
    gen_payload = {
        "businessId": business_id,
        "numberOfQuestions": num_questions,
        "difficulty": difficulty,
        "mode": "Graph" if mode == "Graph-based (CFG)" else "Traditional"
    }
    gen_res = requests.post(generate_endpoint, json=gen_payload, verify=False).json()
    gen_time = int((time.time() - start_time) * 1000)
    
    # 2. Đánh giá Coverage
    print(f"2. Gọi Đánh giá Coverage ({mode})...")
    cov_res = requests.post(f"{api_url}/api/WorkflowAssessment/assess-from-response", json=gen_res, verify=False).json()
    
    # 3. Đánh giá Accuracy
    print(f"3. Gọi Đánh giá Accuracy ({mode})...")
    acc_res = requests.post(f"{api_url}/api/WorkflowAssessment/assess-accuracy", json=gen_res, verify=False).json()
    
    # 4. Đánh giá Difficulty
    print(f"4. Gọi Đánh giá Difficulty ({mode})...")
    diff_res = requests.post(f"{api_url}/api/WorkflowAssessment/assess-difficulty", json=gen_res, verify=False).json()
    
    print(f"Hoàn tất Pipeline {mode}.")
    
    return gen_time, gen_res, cov_res, acc_res, diff_res

def assemble_results(gen_time, gen_res, cov_res, acc_res, diff_res):
    questions_list = gen_res.get("generatedQuestionDtos", gen_res.get("GeneratedQuestionDtos", []))
    
    details = []
    
    for q in questions_list:
        q_text = q.get("question", q.get("Question", ""))
        
        # Tìm coverage
        cov = next((c for c in cov_res.get("questionResults", []) if c.get("question") == q_text), {})
        # Tìm accuracy
        acc = next((a for a in acc_res.get("questionResults", []) if a.get("question") == q_text), {})
        acc_result = acc.get("accuracyResult", {})
        # Tìm difficulty
        diff = next((d for d in diff_res.get("questionResults", []) if d.get("question") == q_text), {})
        diff_result = diff.get("difficultyResult", {})
        
        details.append({
            "question": q_text,
            "coverage": cov.get("coverage", 0),
            "activeNodes": cov.get("activeNodeCount", 0),
            "isAccurate": acc_result.get("isAccurate", False),
            "cyclomatic": diff_result.get("cyclomaticComplexity", 0),
            "evaluationNotes": acc_result.get("finalVerdict", "")
        })
        
    avg_coverage = cov_res.get("averageTotalCoverage", 0)
    
    # Tính trung bình active nodes, accuracy rate, complexity từ các câu hỏi
    avg_active_nodes = sum(d["activeNodes"] for d in details) / len(details) if details else 0
    accuracy_rate = sum(1 for d in details if d["isAccurate"]) / len(details) if details else 0
    avg_complexity = diff_res.get("averageCyclomaticComplexity", sum(diff.get("difficultyResult", {}).get("cyclomaticComplexity", 0) for diff in diff_res.get("questionResults", [])) / len(details) if details else 0)

    return {
        "time": gen_time,
        "inputTokens": gen_res.get("inputTokens", 0),
        "outputTokens": gen_res.get("outputTokens", 0),
        "coverage": avg_coverage,
        "activeNodes": int(avg_active_nodes),
        "accuracy": accuracy_rate,
        "complexity": avg_complexity,
        "details": details,
        "businessName": gen_res.get("businessName", "")
    }

def main():
    parser = argparse.ArgumentParser(description="Run Repo Into Graph Benchmark completely independently via Workflow APIs")
    parser.add_argument("--businessId", type=str, required=True, help="Business ID to test")
    parser.add_argument("--runId", type=str, required=True, help="Run ID to tag the test (e.g. RUN_01)")
    parser.add_argument("--numberOfQuestions", type=int, default=5, help="Number of questions to generate")
    parser.add_argument("--difficulty", type=str, default="Medium", help="Target difficulty")
    parser.add_argument("--apiUrl", type=str, default="https://localhost:55060", help="Base API URL")

    args = parser.parse_args()

    print(f"=== Bắt đầu Benchmark cho Business ID: {args.businessId} ===")
    
    try:
        # Traditional
        trad_time, trad_gen, trad_cov, trad_acc, trad_diff = run_pipeline(args.apiUrl, args.businessId, args.numberOfQuestions, args.difficulty, "Traditional")
        trad_data = assemble_results(trad_time, trad_gen, trad_cov, trad_acc, trad_diff)
        
        # Graph-based
        cfg_time, cfg_gen, cfg_cov, cfg_acc, cfg_diff = run_pipeline(args.apiUrl, args.businessId, args.numberOfQuestions, args.difficulty, "Graph-based (CFG)")
        cfg_data = assemble_results(cfg_time, cfg_gen, cfg_cov, cfg_acc, cfg_diff)
        
        # Lắp ghép final payload
        results = {
            "runId": args.runId,
            "businessName": trad_data["businessName"],
            "difficulty": args.difficulty,
            
            "tradTime": trad_data["time"],
            "tradInputTokens": trad_data["inputTokens"],
            "tradOutputTokens": trad_data["outputTokens"],
            "tradCoverage": trad_data["coverage"],
            "tradActiveNodes": trad_data["activeNodes"],
            "tradAccuracy": trad_data["accuracy"],
            "tradComplexity": trad_data["complexity"],
            "tradDetails": trad_data["details"],
            
            "cfgTime": cfg_data["time"],
            "cfgInputTokens": cfg_data["inputTokens"],
            "cfgOutputTokens": cfg_data["outputTokens"],
            "cfgCoverage": cfg_data["coverage"],
            "cfgActiveNodes": cfg_data["activeNodes"],
            "cfgAccuracy": cfg_data["accuracy"],
            "cfgComplexity": cfg_data["complexity"],
            "cfgDetails": cfg_data["details"]
        }
        
    except requests.exceptions.RequestException as e:
        print(f"LỖI HTTP CALL: {e}")
        return
    except Exception as e:
        print(f"LỖI: {e}")
        return

    excel_path = find_excel_file()
    if not os.path.exists(excel_path):
        print(f"LỖI: Không tìm thấy file excel template tại {excel_path}.")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws_summary = wb["Thử nghiệm"]
        append_summary(ws_summary, results)

        # 2. Update Sheet: Chi tiết Câu hỏi (Details)
        if "Chi tiết Câu hỏi" not in wb.sheetnames:
            ws_details = wb.create_sheet("Chi tiết Câu hỏi")
            ws_details.append([
                "Run ID", "Phương Pháp", "STT Câu", 
                "Nội dung Câu hỏi Nghiệp vụ", "Active Nodes Count", 
                "Coverage (%)", "Độ chính xác (Accuracy)", 
                "Nhánh rẽ Kích hoạt (Gateways)", "Ghi chú Đánh giá"
            ])
        else:
            ws_details = wb["Chi tiết Câu hỏi"]
        append_details(ws_details, results["runId"], "Traditional", results["tradDetails"])
        append_details(ws_details, results["runId"], "Graph-based (CFG)", results["cfgDetails"])

        wb.save(excel_path)
        print(f"\n✅ Đã xuất Data ra Excel thành công: {os.path.basename(excel_path)}")
    except Exception as e:
        print(f"LỖI EXCEL: {e}")

if __name__ == "__main__":
    main()
