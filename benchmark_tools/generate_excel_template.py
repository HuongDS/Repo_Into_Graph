import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Setup styles
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid") # Soft Blue
section_font = Font(name="Calibri", size=11, bold=True, color="1F4E78")

highlight_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Soft Green
highlight_font = Font(name="Calibri", size=11, bold=True, color="375623")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

bold_font = Font(name="Calibri", size=11, bold=True)
regular_font = Font(name="Calibri", size=11)

# ---------------------------------------------------------
# SHEET 1: Summary Dashboard
# ---------------------------------------------------------
ws1 = wb.active
ws1.title = "Báo cáo So sánh (Dashboard)"
ws1.views.sheetView[0].showGridLines = True

# Title
ws1.merge_cells("A1:G1")
ws1["A1"] = "BÁO CÁO KẾT QUẢ THỬ NGHIỆM SO SÁNH: TRADITIONAL (CODE THÔ) VS GRAPH-BASED (CFG)"
ws1["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
ws1["A1"].alignment = Alignment(vertical="center")

ws1.append([])

headers_summary = ["Hạng mục Đánh giá (Metrics)", "Đơn vị", "Phương pháp Code Thô (Traditional)", "Phương pháp Đồ thị (CFG)", "Chênh lệch (Delta)", "% Tối ưu / Tăng trưởng", "Ghi chú & Nhận xét"]
ws1.append(headers_summary)

summary_data = [
    ["Độ bao phủ trung bình (Average Total Coverage)", "%", "='Thử nghiệm'!H25", "='Thử nghiệm'!H26", "=D4-C4", "=IF(C4=0,0,(D4-C4)/C4)", "Chỉ số từ API assess-from-response"],
    ["Độ bao phủ theo Workflow (Coverage Workflow/Global)", "%", 0.45, 0.82, "=D5-C5", "=IF(C5=0,0,(D5-C5)/C5)", "Tỷ lệ nút workflow được chạm đến"],
    ["Độ chính xác trung bình (Average Accuracy Rate)", "%", "='Thử nghiệm'!J25", "='Thử nghiệm'!J26", "=D6-C6", "=IF(C6=0,0,(D6-C6)/C6)", "Chỉ số từ API assess-accuracy"],
    ["Thời gian sinh câu hỏi trung bình (Avg Gen Time)", "ms", "='Thử nghiệm'!E25", "='Thử nghiệm'!E26", "=D7-C7", "=IF(C7=0,0,(D7-C7)/C7)", "Thời gian Postman nhận Response"],
    ["Tổng Token đầu vào (Total Input Tokens)", "tokens", "='Thử nghiệm'!F25", "='Thử nghiệm'!F26", "=D8-C8", "=IF(C8=0,0,(D8-C8)/C8)", "Token truyền vào Gemini AI"],
    ["Tổng Token đầu ra (Total Output Tokens)", "tokens", "='Thử nghiệm'!G25", "='Thử nghiệm'!G26", "=D9-C9", "=IF(C9=0,0,(D9-C9)/C9)", "Token Gemini AI phản hồi"],
    ["Số câu hỏi hợp lệ / Đúng logic", "câu", 12, 19, "=D10-C10", "=IF(C10=0,0,(D10-C10)/C10)", "Số câu không vi phạm logic code"],
    ["Độ phức tạp Cyclomatic Avg (Số cạnh active)", "cạnh", 1.8, 3.4, "=D11-C11", "=IF(C11=0,0,(D11-C11)/C11)", "Chỉ số từ API assess-difficulty"]
]

for row in summary_data:
    ws1.append(row)

# Styling Sheet 1 Header
for col_idx in range(1, 8):
    cell = ws1.cell(row=3, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Styling Sheet 1 Data
for row_idx in range(4, 12):
    for col_idx in range(1, 8):
        cell = ws1.cell(row=row_idx, column=col_idx)
        cell.font = regular_font
        cell.border = thin_border
        if col_idx in [3, 4, 5]:
            if row_idx in [4, 5, 6]:
                cell.number_format = '0.00%'
            elif row_idx in [7, 8, 9]:
                cell.number_format = '#,##0'
            else:
                cell.number_format = '0.0'
        elif col_idx == 6:
            cell.number_format = '+0.00%;-0.00%;0.00%'
            cell.font = bold_font

# ---------------------------------------------------------
# SHEET 2: Raw Test Data
# ---------------------------------------------------------
ws2 = wb.create_sheet(title="Thử nghiệm")
ws2.views.sheetView[0].showGridLines = True

ws2["A1"] = "DANH SÁCH BẢNG GHI THỬ NGHIỆM TRÊN POSTMAN (RUN LOG)"
ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")

headers_test = [
    "STT", "Mã Kịch Bản (Run ID)", "Tên Business / Flow", "Phương Pháp (Mode)", 
    "Thời gian Gen (ms)", "Input Tokens", "Output Tokens", "Coverage Total (%)", 
    "Active Nodes Count", "Accuracy Rate (%)", "Số Cạnh Active (Cyclomatic)", "Độ Khó Đặt Ra", "Ghi Chú"
]
ws2.append([])
ws2.append(headers_test)

# Add dummy sample rows
sample_runs = [
    [1, "RUN_01_TRAD", "Đăng ký tài khoản", "Traditional", 2450, 1200, 350, 0.42, 3, 0.70, 2, "Medium", "Code thô bỏ sót bước validate email"],
    [2, "RUN_01_CFG", "Đăng ký tài khoản", "Graph-based (CFG)", 2100, 850, 380, 0.78, 6, 0.95, 4, "Medium", "Phủ kín các nhánh exception"],
    [3, "RUN_02_TRAD", "Thanh toán đơn hàng", "Traditional", 3100, 1800, 420, 0.38, 4, 0.65, 2, "Hard", "Thiếu luồng rollback giao dịch"],
    [4, "RUN_02_CFG", "Thanh toán đơn hàng", "Graph-based (CFG)", 2750, 1100, 490, 0.85, 9, 0.92, 5, "Hard", "Phủ đầy đủ các gateway điều kiện"],
    [5, "RUN_03_TRAD", "Duyệt hồ sơ vay", "Traditional", 2900, 1550, 400, 0.50, 5, 0.75, 3, "Medium", "Chỉ hỏi luồng thành công"],
    [6, "RUN_03_CFG", "Duyệt hồ sơ vay", "Graph-based (CFG)", 2300, 950, 460, 0.88, 8, 0.98, 4, "Medium", "Phủ chính xác các bước phê duyệt"]
]

for row in sample_runs:
    ws2.append(row)

# Fill empty rows up to row 24 for manual entry
for i in range(7, 23):
    ws2.append([i-4, f"RUN_{i-4:02d}", "", "", "", "", "", "", "", "", "", "", ""])

# Add Summary rows at row 25 & 26
ws2.cell(row=25, column=3, value="TRUNG BÌNH TRADITIONAL").font = bold_font
ws2.cell(row=25, column=4, value="Traditional").font = bold_font
ws2.cell(row=25, column=5, value="=AVERAGEIF(D3:D24,\"Traditional\",E3:E24)")
ws2.cell(row=25, column=6, value="=SUMIF(D3:D24,\"Traditional\",F3:F24)")
ws2.cell(row=25, column=7, value="=SUMIF(D3:D24,\"Traditional\",G3:G24)")
ws2.cell(row=25, column=8, value="=AVERAGEIF(D3:D24,\"Traditional\",H3:H24)")
ws2.cell(row=25, column=9, value="=AVERAGEIF(D3:D24,\"Traditional\",I3:I24)")
ws2.cell(row=25, column=10, value="=AVERAGEIF(D3:D24,\"Traditional\",J3:J24)")
ws2.cell(row=25, column=11, value="=AVERAGEIF(D3:D24,\"Traditional\",K3:K24)")

ws2.cell(row=26, column=3, value="TRUNG BÌNH GRAPH-BASED").font = bold_font
ws2.cell(row=26, column=4, value="Graph-based (CFG)").font = bold_font
ws2.cell(row=26, column=5, value="=AVERAGEIF(D3:D24,\"Graph-based (CFG)\",E3:E24)")
ws2.cell(row=26, column=6, value="=SUMIF(D3:D24,\"Graph-based (CFG)\",F3:F24)")
ws2.cell(row=26, column=7, value="=SUMIF(D3:D24,\"Graph-based (CFG)\",G3:G24)")
ws2.cell(row=26, column=8, value="=AVERAGEIF(D3:D24,\"Graph-based (CFG)\",H3:H24)")
ws2.cell(row=26, column=9, value="=AVERAGEIF(D3:D24,\"Graph-based (CFG)\",I3:I24)")
ws2.cell(row=26, column=10, value="=AVERAGEIF(D3:D24,\"Graph-based (CFG)\",J3:J24)")
ws2.cell(row=26, column=11, value="=AVERAGEIF(D3:D24,\"Graph-based (CFG)\",K3:K24)")

# Styling Sheet 2 Header
for col_idx in range(1, 14):
    cell = ws2.cell(row=3, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Styling Sheet 2 Data
for row_idx in range(4, 27):
    for col_idx in range(1, 14):
        cell = ws2.cell(row=row_idx, column=col_idx)
        cell.font = regular_font
        cell.border = thin_border
        if col_idx in [8, 10]:
            cell.number_format = '0.00%'
        elif col_idx in [5, 6, 7]:
            cell.number_format = '#,##0'
        elif col_idx in [9, 11]:
            cell.number_format = '0.0'
        
        if row_idx in [25, 26]:
            cell.fill = section_fill
            cell.font = bold_font

# ---------------------------------------------------------
# SHEET 3: Question Level Details
# ---------------------------------------------------------
ws3 = wb.create_sheet(title="Chi tiết Câu hỏi")
ws3.views.sheetView[0].showGridLines = True

ws3["A1"] = "CHI TIẾT ĐÁNH GIÁ TỪNG CÂU HỎI TRONG BỘ TEST (QUESTION LEVEL LOG)"
ws3["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")

headers_q = [
    "Run ID", "Phương Pháp", "STT Câu", "Nội dung Câu hỏi Nghiệp vụ", 
    "Active Nodes Count", "Coverage (%)", "Độ chính xác (Accuracy)", "Nhánh rẽ Kích hoạt (Gateways)", "Ghi chú Đánh giá"
]
ws3.append([])
ws3.append(headers_q)

sample_q = [
    ["RUN_01_TRAD", "Traditional", 1, "Khi người dùng nhập sai mật khẩu 3 lần thì hệ thống xử lý thế nào?", 2, 0.40, "Đúng", "1 Gateway", "Chỉ phủ bước kiểm tra password"],
    ["RUN_01_CFG", "Graph-based (CFG)", 1, "Khi người dùng nhập sai mật khẩu 3 lần liên tiếp, tài khoản có bị khóa và gửi email cảnh báo không?", 5, 0.85, "Đúng", "3 Gateways", "Phủ đầy đủ luồng khóa tài khoản & gửi email warning"],
    ["RUN_02_TRAD", "Traditional", 1, "Luồng thanh toán MoMo thực hiện qua các bước nào?", 3, 0.35, "Sai", "0 Gateway", "Bỏ qua bước verify IPN checksum từ MoMo"],
    ["RUN_02_CFG", "Graph-based (CFG)", 1, "Trong luồng thanh toán MoMo, nếu chữ ký checksum IPN không hợp lệ thì giao dịch bị hủy thế nào?", 7, 0.90, "Đúng", "4 Gateways", "Bắt chính xác hàm VerifySignature và RollbackTransaction"]
]

for row in sample_q:
    ws3.append(row)

for i in range(5, 30):
    ws3.append(["", "", i-4, "", "", "", "", "", ""])

# Style Sheet 3
for col_idx in range(1, 10):
    cell = ws3.cell(row=3, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx in range(4, 30):
    for col_idx in range(1, 10):
        cell = ws3.cell(row=row_idx, column=col_idx)
        cell.font = regular_font
        cell.border = thin_border
        if col_idx == 6:
            cell.number_format = '0.00%'

# ---------------------------------------------------------
# SHEET 4: Postman Guide
# ---------------------------------------------------------
ws4 = wb.create_sheet(title="Hướng dẫn Test Postman")
ws4.views.sheetView[0].showGridLines = True

ws4["A1"] = "HƯỚNG DẪN CHẠY TEST TỰ ĐỘNG BẰNG POSTMAN & LẤY SỐ LIỆU"
ws4["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E78")

guide_content = [
    ["Bước", "Thao tác trên Postman", "API Endpoint", "Số liệu thu thập & Điền vào Excel"],
    ["1", "Mở Postman Collection 'Repo Into Graph - A/B Testing Automation'", "-", "Đảm bảo đã set biến {{baseUrl}} và {{businessId}}"],
    ["2", "Chạy folder 'Phương pháp Cũ (Raw Code)'", "POST /api/QuestionGenerator/generate-traditional", "Lấy Thời gian Gen (ms), Input Tokens, Output Tokens -> Điền vào cột E, F, G sheet 'Thử nghiệm'"],
    ["3", "Chạy API Assess Coverage (Traditional)", "POST /api/WorkflowAssessment/assess-from-response", "Lấy giá trị 'averageTotalCoverage' -> Điền vào cột Coverage Total (cột H)"],
    ["4", "Chạy API Assess Accuracy (Traditional)", "POST /api/WorkflowAssessment/assess-accuracy", "Lấy tỷ lệ câu hỏi Valid / Total -> Điền vào cột Accuracy Rate (cột J)"],
    ["5", "Chạy API Assess Difficulty (Traditional)", "POST /api/WorkflowAssessment/assess-difficulty", "Lấy 'totalEdgesInSubgraph' trung bình -> Điền vào cột K"],
    ["6", "Chạy folder 'Phương pháp Mới (CFG / Graph-based)'", "POST /api/QuestionGenerator/generate-graph", "Lặp lại bước 2-5 cho phương pháp CFG để so sánh đối chứng"],
    ["7", "Xem kết quả tự động trên Dashboard", "Sheet 'Báo cáo So sánh (Dashboard)'", "File Excel tự động tính toán % Tăng trưởng (% Delta) của CFG so với Traditional"]
]

for row in guide_content:
    ws4.append(row)

for col_idx in range(1, 5):
    cell = ws4.cell(row=2, column=col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row_idx in range(3, 11):
    for col_idx in range(1, 5):
        cell = ws4.cell(row=row_idx, column=col_idx)
        cell.font = regular_font
        cell.border = thin_border

# Auto adjust column widths for all sheets
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.number_format and '%' in cell.number_format:
                val_str += '   '
            max_len = max(max_len, len(val_str))
        sheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

output_path = r"d:\TaiLieu_SU26\DUAL_CAPSTONE\Repo_Into_Graph_Solutions\CFG_vs_Traditional_Benchmark_Template.xlsx"
wb.save(output_path)
print(f"Successfully generated template at: {output_path}")
