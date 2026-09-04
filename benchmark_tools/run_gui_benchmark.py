import os
import time
import json
import threading
import datetime
import urllib3
import requests
import openpyxl
import customtkinter as ctk
from tkinter import filedialog, messagebox

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- THEME CONFIGURATION ---
ctk.set_appearance_mode("Light")  # The user explicitly requested Light mode
ctk.set_default_color_theme("blue")

class BenchmarkApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        self.title("Repo Into Graph - Auto Benchmark & Testing Tool")
        self.geometry("850x800")
        self.resizable(True, True)
        self.configure(fg_color="#F0F4F8")  # Soft light background

        
        self.loaded_businesses = []  # List of dicts: {"name": "...", "id": "..."}
        
        # --- UI SETUP ---
        self.setup_ui()
        
    def setup_ui(self):
        # 1. HEADER
        header_frame = ctk.CTkFrame(self, fg_color="transparent")
        header_frame.pack(pady=(20, 10), fill="x")
        header_label = ctk.CTkLabel(header_frame, text="📊 AUTO BENCHMARK & TESTING TOOL", font=ctk.CTkFont(size=26, weight="bold", family="Segoe UI"))
        header_label.pack()
        sub_label = ctk.CTkLabel(header_frame, text="Repo Into Graph Solutions", font=ctk.CTkFont(size=14, slant="italic", family="Segoe UI"), text_color="gray50")
        sub_label.pack()
        
        # 2. FRAME 1: Data Loading (Nạp Dữ liệu)
        frame_data = ctk.CTkFrame(self, corner_radius=12, fg_color="#FFFFFF", border_width=1, border_color="#E1E5EB")
        frame_data.pack(padx=30, pady=10, fill="x")
        
        lbl_data = ctk.CTkLabel(frame_data, text="📝 BƯỚC 1: NẠP TẬP DỮ LIỆU ĐẦU VÀO", font=ctk.CTkFont(size=16, weight="bold"))
        lbl_data.grid(row=0, column=0, padx=20, pady=15, sticky="w")
        
        btn_generate_template = ctk.CTkButton(frame_data, text="📑 Tạo Template Mẫu", command=self.generate_template, fg_color="#107C41", hover_color="#0b5c30", font=ctk.CTkFont(weight="bold"))
        btn_generate_template.grid(row=0, column=1, padx=10, pady=15)
        
        btn_load_excel = ctk.CTkButton(frame_data, text="📂 Chọn File Excel", command=self.load_excel, fg_color="#2b579a", hover_color="#1d3d6b", font=ctk.CTkFont(weight="bold"))
        btn_load_excel.grid(row=0, column=2, padx=10, pady=15)
        
        self.lbl_loaded_status = ctk.CTkLabel(frame_data, text="⚠ Chưa nạp file danh sách nghiệp vụ nào.", text_color="#d13438", font=ctk.CTkFont(slant="italic"))
        self.lbl_loaded_status.grid(row=1, column=0, columnspan=3, padx=20, pady=(0, 15), sticky="w")
        
        # 3. FRAME 2: Configuration (Cấu hình Chạy Test)
        frame_config = ctk.CTkFrame(self, corner_radius=12, fg_color="#FFFFFF", border_width=1, border_color="#E1E5EB")
        frame_config.pack(padx=30, pady=10, fill="x")
        
        lbl_config = ctk.CTkLabel(frame_config, text="⚙️ BƯỚC 2: CẤU HÌNH KIỂM THỬ", font=ctk.CTkFont(size=16, weight="bold"))
        lbl_config.grid(row=0, column=0, padx=20, pady=15, sticky="w", columnspan=2)
        
        # - Chọn nghiệp vụ
        ctk.CTkLabel(frame_config, text="📌 Chọn Nghiệp Vụ:", font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, padx=20, pady=10, sticky="w")
        self.combo_business = ctk.CTkComboBox(frame_config, values=["(Hãy nạp file Excel trước)"], width=450, state="readonly")
        self.combo_business.grid(row=1, column=1, padx=10, pady=10, sticky="w", columnspan=3)
        self.combo_business.set("(Hãy nạp file Excel trước)")
        
        # - Phương pháp
        ctk.CTkLabel(frame_config, text="🔬 Phương Pháp Test:", font=ctk.CTkFont(weight="bold")).grid(row=2, column=0, padx=20, pady=10, sticky="w")
        self.radio_var = ctk.StringVar(value="Both")
        r_both = ctk.CTkRadioButton(frame_config, text="A/B Test (Cả Hai)", variable=self.radio_var, value="Both")
        r_trad = ctk.CTkRadioButton(frame_config, text="Truyền Thống", variable=self.radio_var, value="Traditional")
        r_cfg = ctk.CTkRadioButton(frame_config, text="CFG (Graph)", variable=self.radio_var, value="CFG")
        
        r_both.grid(row=2, column=1, padx=10, pady=10, sticky="w")
        r_trad.grid(row=2, column=2, padx=10, pady=10, sticky="w")
        r_cfg.grid(row=2, column=3, padx=10, pady=10, sticky="w")
        
        # - Số câu hỏi & Độ khó
        ctk.CTkLabel(frame_config, text="🔢 Số Lượng Câu Hỏi:", font=ctk.CTkFont(weight="bold")).grid(row=3, column=0, padx=20, pady=10, sticky="w")
        self.combo_num = ctk.CTkComboBox(frame_config, values=["5", "10", "15", "20"], width=120, state="readonly")
        self.combo_num.grid(row=3, column=1, padx=10, pady=10, sticky="w")
        self.combo_num.set("5")
        
        ctk.CTkLabel(frame_config, text="🔥 Độ Khó:", font=ctk.CTkFont(weight="bold")).grid(row=3, column=2, padx=10, pady=10, sticky="e")
        self.combo_difficulty = ctk.CTkComboBox(frame_config, values=["Easy", "Medium", "Hard"], width=120, state="readonly")
        self.combo_difficulty.set("Medium")
        self.combo_difficulty.grid(row=3, column=3, padx=10, pady=10, sticky="w")
        
        # 4. FRAME 3: Execution & Logs
        frame_exec = ctk.CTkFrame(self, corner_radius=12, fg_color="#FFFFFF", border_width=1, border_color="#E1E5EB")
        frame_exec.pack(padx=30, pady=10, fill="both", expand=True)
        
        button_frame = ctk.CTkFrame(frame_exec, fg_color="transparent")
        button_frame.pack(fill="x", padx=20, pady=15)
        
        self.btn_run = ctk.CTkButton(
            button_frame, 
            text="🚀 CHẠY BENCHMARK", 
            font=ctk.CTkFont(size=18, weight="bold"), 
            height=50, 
            corner_radius=8,
            fg_color="#0066CC", 
            hover_color="#0052A3",
            command=self.start_benchmark_thread
        )
        self.btn_run.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        self.btn_view_logs = ctk.CTkButton(
            button_frame, 
            text="📂 XEM LOGS", 
            font=ctk.CTkFont(size=18, weight="bold"), 
            height=50, 
            corner_radius=8,
            fg_color="#107C41", 
            hover_color="#0b5c30",
            command=self.open_logs_viewer
        )
        self.btn_view_logs.pack(side="right", fill="x", expand=True, padx=(10, 0))
        self.lbl_status = ctk.CTkLabel(frame_exec, text="Trạng thái: Sẵn sàng", font=ctk.CTkFont(weight="bold", slant="italic"), text_color="#0066CC")
        self.lbl_status.pack(anchor="w", padx=20)

        self.progress_bar = ctk.CTkProgressBar(frame_exec, height=12, corner_radius=6, progress_color="#107C41", fg_color="#E1E5EB")
        self.progress_bar.pack(fill="x", padx=20, pady=(5, 15))
        self.progress_bar.set(0)
        
        self.textbox_log = ctk.CTkTextbox(frame_exec, height=200, state="disabled", font=ctk.CTkFont(family="Consolas", size=13), fg_color="#F8F9FA", text_color="#333333", border_width=1, border_color="#E1E5EB")
        self.textbox_log.pack(padx=20, pady=(0, 20), fill="both", expand=True)
    # --- ACTIONS ---
    
    def set_status(self, text):
        self.lbl_status.configure(text=text)
        self.update_idletasks()
        
    def set_progress(self, value):
        self.progress_bar.set(value)
        self.update_idletasks()
        
    def log(self, message):
        """Helper to append log messages thread-safely."""
        self.textbox_log.configure(state="normal")
        self.textbox_log.insert("end", message + "\n")
        self.textbox_log.see("end")
        self.textbox_log.configure(state="disabled")
        self.update_idletasks()

    def generate_template(self):
        try:
            filename = "DanhSachNghiepVu.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Nghiệp Vụ"
            
            # Header
            ws.cell(row=1, column=1, value="Business Name")
            ws.cell(row=1, column=2, value="Business ID")
            
            # Data
            ws.cell(row=2, column=1, value="Đăng nhập hệ thống")
            ws.cell(row=2, column=2, value="12345678-1234-1234-1234-123456789012")
            
            ws.cell(row=3, column=1, value="Thanh toán giỏ hàng")
            ws.cell(row=3, column=2, value="87654321-4321-4321-4321-210987654321")
            
            wb.save(filename)
            messagebox.showinfo("Thành công", f"Đã tạo file mẫu '{filename}' tại thư mục hiện tại.\nHãy mở file lên và thay thế bằng ID thật của bạn.")
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể tạo file: {str(e)}")

    def load_excel(self):
        filepath = filedialog.askopenfilename(title="Chọn file Excel", filetypes=[("Excel files", "*.xlsx *.xls")])
        if not filepath:
            return
            
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            
            # Find column indices for Business Name and Business ID
            name_col = None
            id_col = None
            
            for col in range(1, ws.max_column + 1):
                header = str(ws.cell(row=1, column=col).value).strip()
                if header == "Business Name":
                    name_col = col
                elif header == "Business ID":
                    id_col = col
                    
            if name_col is None or id_col is None:
                messagebox.showerror("Lỗi Cấu Trúc", "File Excel phải có ít nhất 2 cột: 'Business Name' và 'Business ID'")
                return
                
            self.loaded_businesses = []
            combo_values = []
            
            for row in range(2, ws.max_row + 1):
                b_name = str(ws.cell(row=row, column=name_col).value).strip()
                b_id = str(ws.cell(row=row, column=id_col).value).strip()
                
                if b_id and b_id.lower() not in ["none", "nan", ""]:
                    self.loaded_businesses.append({"name": b_name, "id": b_id})
                    combo_values.append(f"{b_name} ({b_id})")
                    
            if not self.loaded_businesses:
                self.lbl_loaded_status.configure(text="Không tìm thấy dữ liệu hợp lệ trong file.", text_color="red")
                return
                
            self.combo_business.configure(values=combo_values)
            self.combo_business.set(combo_values[0])
            self.lbl_loaded_status.configure(text=f"✅ Đã tải thành công {len(self.loaded_businesses)} nghiệp vụ.", text_color="green")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Đã xảy ra lỗi khi đọc file:\n{str(e)}")

    def get_selected_business_id(self):
        selected_text = self.combo_business.get()
        for b in self.loaded_businesses:
            if b["id"] in selected_text:
                return b["id"], b["name"]
        return None, None

    def start_benchmark_thread(self):
        if not self.loaded_businesses:
            messagebox.showwarning("Cảnh báo", "Vui lòng nạp danh sách nghiệp vụ trước!")
            return
            
        b_id, b_name = self.get_selected_business_id()
        if not b_id:
            messagebox.showwarning("Cảnh báo", "Nghiệp vụ đã chọn không hợp lệ!")
            return
            
        answer = messagebox.askyesnocancel("Ghi dữ liệu Excel", "Bạn muốn GHI TIẾP vào dữ liệu hiện có (Yes)\nHay XÓA TOÀN BỘ dữ liệu cũ (No)?\n\n(Bấm Cancel để hủy chạy)")
        if answer is None:
            return
            
        clear_old_data = not answer  # True if No (Clear), False if Yes (Append)
            
        self.btn_run.configure(state="disabled")
        self.textbox_log.configure(state="normal")
        self.textbox_log.delete("1.0", "end")
        self.textbox_log.configure(state="disabled")
        
        # Start thread
        thread = threading.Thread(target=self.run_benchmark_workflow, args=(b_id, b_name, clear_old_data))
        thread.start()

    # --- BENCHMARK WORKFLOW LOGIC ---
    
    def run_pipeline(self, api_url, business_id, num_questions, difficulty, mode):
        self.log(f"\n--- Bắt đầu Pipeline {mode} ---")
        self.set_status(f"Trạng thái: Khởi động Pipeline {mode}...")
        self.set_progress(0.1)
        
        generate_endpoint = f"{api_url}/api/QuestionGenerator/generate-traditional" if mode == "Traditional" else f"{api_url}/api/QuestionGenerator/generate-graph"
        
        self.set_status(f"Trạng thái: Đang sinh {num_questions} câu hỏi ({mode})...")
        self.log(f"1. Calling QuestionGenerator ({mode})...")
        start_time = time.time()
        gen_payload = {
            "businessId": business_id,
            "numberOfQuestions": num_questions,
            "difficulty": difficulty,
            "mode": "Graph" if mode == "Graph-based (CFG)" else "Traditional"
        }
        
        gen_res = requests.post(generate_endpoint, json=gen_payload, verify=False).json()
        gen_time = int((time.time() - start_time) * 1000)
        
        questions = gen_res.get("generatedQuestionDtos", gen_res.get("GeneratedQuestionDtos", []))
        self.log(f"   -> Đã sinh thành công {len(questions)} câu hỏi trong {gen_time}ms.")
        self.set_progress(0.3)
        
        self.set_status(f"Trạng thái: Đang chấm điểm Coverage ({mode})...")
        self.log(f"2. Calling Coverage Assessment ({mode})...")
        cov_res = requests.post(f"{api_url}/api/WorkflowAssessment/assess-from-response", json=gen_res, verify=False).json()
        self.log(f"   -> Hoàn thành Coverage Assessment.")
        self.set_progress(0.5)
        
        self.set_status(f"Trạng thái: Đang chấm điểm Accuracy ({mode})...")
        self.log(f"3. Calling Accuracy Assessment ({mode})...")
        acc_res = requests.post(f"{api_url}/api/WorkflowAssessment/assess-accuracy", json=gen_res, verify=False).json()
        self.log(f"   -> Hoàn thành Accuracy Assessment.")
        self.set_progress(0.7)
        
        self.set_status(f"Trạng thái: Đang đánh giá Độ khó ({mode})...")
        self.log(f"4. Calling Difficulty Assessment ({mode})...")
        diff_res = requests.post(f"{api_url}/api/WorkflowAssessment/assess-difficulty", json=gen_res, verify=False).json()
        self.log(f"   -> Hoàn thành Difficulty Assessment.")
        self.set_progress(0.9)
        
        self.log(f"Hoàn tất Pipeline {mode}.")
        self.set_status(f"Trạng thái: Đã hoàn tất Pipeline {mode}.")
        return gen_time, gen_res, cov_res, acc_res, diff_res

    def assemble_results(self, gen_time, gen_res, cov_res, acc_res, diff_res):
        questions_list = gen_res.get("generatedQuestionDtos", gen_res.get("GeneratedQuestionDtos", []))
        details = []
        
        for q in questions_list:
            q_text = q.get("question", q.get("Question", ""))
            cov = next((c for c in cov_res.get("questionResults", []) if c.get("question") == q_text), {})
            acc = next((a for a in acc_res.get("questionResults", []) if a.get("question") == q_text), {})
            acc_result = acc.get("accuracyResult", {})
            diff = next((d for d in diff_res.get("questionResults", []) if d.get("question") == q_text), {})
            diff_result = diff.get("difficultyResult", {})
            
            details.append({
                "question": q_text,
                "coverage": cov.get("coverage", 0),
                "activeNodes": cov.get("activeNodeCount", 0),
                "isAccurate": acc_result.get("isAccurate", False),
                "cyclomatic": diff_result.get("cyclomaticComplexity", 0),
                "evaluationNotes": acc_result.get("finalVerdict", "")
            })
            
        avg_coverage = cov_res.get("averageTotalCoverage", 0)
        avg_active_nodes = sum(d["activeNodes"] for d in details) / len(details) if details else 0
        accuracy_rate = sum(1 for d in details if d["isAccurate"]) / len(details) if details else 0
        avg_complexity = diff_res.get("averageCyclomaticComplexity", sum(diff.get("difficultyResult", {}).get("cyclomaticComplexity", 0) for diff in diff_res.get("questionResults", [])) / len(details) if details else 0)

        return {
            "time": gen_time,
            "inputTokens": gen_res.get("inputTokens", 0),
            "outputTokens": gen_res.get("outputTokens", 0),
            "coverage": avg_coverage,
            "activeNodes": int(avg_active_nodes),
            "accuracy": accuracy_rate,
            "complexity": avg_complexity,
            "details": details,
            "businessName": gen_res.get("businessName", "")
        }

    def run_benchmark_workflow(self, business_id, business_name, clear_old_data=False):
        method = self.radio_var.get()
        num_questions = int(self.combo_num.get())
        difficulty = self.combo_difficulty.get()
        api_url = "https://localhost:55060"
        
        # Auto-generate Run ID based on timestamp
        run_id = datetime.datetime.now().strftime("RUN_%d%m_%H%M")
        
        self.log(f"=== BẮT ĐẦU BENCHMARK ===")
        self.log(f"Nghiệp vụ: {business_name}")
        self.log(f"Phương pháp: {method}")
        self.log(f"Mã Run ID tự động: {run_id}")
        if clear_old_data:
            self.log(f"[CHÚ Ý] Sẽ XÓA TOÀN BỘ dữ liệu cũ trước khi ghi.")
            
        self.set_progress(0.0)
        self.set_status("Trạng thái: Chuẩn bị chạy...")
        
        results = {
            "runId": run_id,
            "businessName": business_name,
            "difficulty": difficulty
        }
        
        raw_logs = {
            "runId": run_id,
            "businessName": business_name,
            "timestamp": datetime.datetime.now().isoformat(),
            "method": method,
            "difficulty": difficulty,
            "requestedQuestions": num_questions,
            "traditional": None,
            "cfg": None
        }
        
        try:
            if method in ["Both", "Traditional"]:
                t_time, t_gen, t_cov, t_acc, t_diff = self.run_pipeline(api_url, business_id, num_questions, difficulty, "Traditional")
                t_data = self.assemble_results(t_time, t_gen, t_cov, t_acc, t_diff)
                
                results["tradTime"] = t_data["time"]
                results["tradInputTokens"] = t_data["inputTokens"]
                results["tradOutputTokens"] = t_data["outputTokens"]
                results["tradCoverage"] = t_data["coverage"]
                results["tradActiveNodes"] = t_data["activeNodes"]
                results["tradAccuracy"] = t_data["accuracy"]
                results["tradComplexity"] = t_data["complexity"]
                results["tradDetails"] = t_data["details"]
                
                raw_logs["traditional"] = {
                    "generate": t_gen,
                    "coverage": t_cov,
                    "accuracy": t_acc,
                    "difficulty": t_diff
                }

            if method in ["Both", "CFG"]:
                c_time, c_gen, c_cov, c_acc, c_diff = self.run_pipeline(api_url, business_id, num_questions, difficulty, "Graph-based (CFG)")
                c_data = self.assemble_results(c_time, c_gen, c_cov, c_acc, c_diff)
                
                results["cfgTime"] = c_data["time"]
                results["cfgInputTokens"] = c_data["inputTokens"]
                results["cfgOutputTokens"] = c_data["outputTokens"]
                results["cfgCoverage"] = c_data["coverage"]
                results["cfgActiveNodes"] = c_data["activeNodes"]
                results["cfgAccuracy"] = c_data["accuracy"]
                results["cfgComplexity"] = c_data["complexity"]
                results["cfgDetails"] = c_data["details"]
                
                raw_logs["cfg"] = {
                    "generate": c_gen,
                    "coverage": c_cov,
                    "accuracy": c_acc,
                    "difficulty": c_diff
                }
                
            # DUMP RAW LOGS TO JSON
            log_dir = "benchmark_logs"
            if not os.path.exists(log_dir):
                os.makedirs(log_dir)
            log_filename = os.path.join(log_dir, f"{run_id}_Details.json")
            with open(log_filename, "w", encoding="utf-8") as f:
                json.dump(raw_logs, f, ensure_ascii=False, indent=4)
            self.log(f"\nĐã lưu log chi tiết vào: {log_filename}")
                
        except Exception as e:
            self.log(f"\n[LỖI API] Đã có lỗi xảy ra: {str(e)}")
            self.set_status("Trạng thái: LỖI")
            self.btn_run.configure(state="normal")
            return

        self.set_progress(1.0)
        self.set_status("Trạng thái: Đang ghi dữ liệu vào Excel...")
        self.log("\nTiến hành ghi dữ liệu vào Excel...")
        self.save_to_excel(results, method, clear_old_data)
        
        self.btn_run.configure(state="normal")
        self.set_status("Trạng thái: HOÀN TẤT!")
        self.log("\n✅ BENCHMARK HOÀN TẤT THÀNH CÔNG!")
        
    def save_to_excel(self, results, method, clear_old_data=False):
        excel_path = "CFG_vs_Traditional_Benchmark_Template.xlsx"
        
        # Try finding it in parent directories just in case
        if not os.path.exists(excel_path):
            current = os.path.abspath(os.path.curdir)
            while current and current != os.path.dirname(current):
                candidate = os.path.join(current, "CFG_vs_Traditional_Benchmark_Template.xlsx")
                if os.path.exists(candidate):
                    excel_path = candidate
                    break
                current = os.path.dirname(current)

        try:
            if not os.path.exists(excel_path):
                self.log(f"[THÔNG BÁO] Không tìm thấy file excel {excel_path}. Đang tự động tạo file mới...")
                wb = openpyxl.Workbook()
                # Initialize the 'Báo cáo' sheet
                ws_report = wb.active
                ws_report.title = "Báo cáo"
                ws_report.append(['BÁO CÁO KẾT QUẢ THỬ NGHIỆM SO SÁNH: TRADITIONAL (CODE THÔ) VS GRAPH-BASED (CFG)'])
                ws_report.append([])
                ws_report.append(['Hạng mục Đánh giá (Metrics)', 'Đơn vị', 'Phương pháp Code Thô (Traditional)', 'Phương pháp Đồ thị (CFG)', 'Chênh lệch (Delta)', '% Tối ưu / Tăng trưởng', 'Ghi chú & Nhận xét'])
                ws_report.append(['Độ bao phủ trung bình (Average Total Coverage)', '%', '', '', '=D4-C4', '=IF(C4=0,0,(D4-C4)/C4)', 'Chỉ số từ API assess-from-response'])
                ws_report.append(['Độ bao phủ theo Workflow (Coverage Workflow/Global)', '%', '', '', '=D5-C5', '=IF(C5=0,0,(D5-C5)/C5)', 'Tỷ lệ nút workflow được chạm đến'])
                ws_report.append(['Độ chính xác trung bình (Average Accuracy Rate)', '%', '', '', '=D6-C6', '=IF(C6=0,0,(D6-C6)/C6)', 'Chỉ số từ API assess-accuracy'])
                ws_report.append(['Thời gian sinh câu hỏi trung bình (Avg Gen Time)', 'ms', '', '', '=D7-C7', '=IF(C7=0,0,(D7-C7)/C7)', 'Thời gian Postman nhận Response'])
                ws_report.append(['Tổng Token đầu vào (Total Input Tokens)', 'tokens', '', '', '=D8-C8', '=IF(C8=0,0,(D8-C8)/C8)', 'Token truyền vào Gemini AI'])
                ws_report.append(['Tổng Token đầu ra (Total Output Tokens)', 'tokens', '', '', '=D9-C9', '=IF(C9=0,0,(D9-C9)/C9)', 'Token Gemini AI phản hồi'])
                ws_report.append(['Số câu hỏi hợp lệ / Đúng logic', 'câu', '', '', '=D10-C10', '=IF(C10=0,0,(D10-C10)/C10)', 'Số câu không vi phạm logic code'])
                ws_report.append(['Độ phức tạp Cyclomatic Avg (Số cạnh active)', 'cạnh', '', '', '=D11-C11', '=IF(C11=0,0,(D11-C11)/C11)', 'Chỉ số từ API assess-difficulty'])
            else:
                wb = openpyxl.load_workbook(excel_path)
            
            # --- RENAME SHEETS IF NECESSARY ---
            if "Báo cáo So sánh (Dashboard)" in wb.sheetnames:
                wb["Báo cáo So sánh (Dashboard)"].title = "Báo cáo"
            if "Thử nghiệm" in wb.sheetnames:
                wb["Thử nghiệm"].title = "Run"
                
            # --- UPDATE BÁO CÁO FORMULAS ---
            if "Báo cáo" in wb.sheetnames:
                ws_report = wb["Báo cáo"]
                # Traditional (Column C)
                ws_report["C4"] = '=AVERAGEIFS(Run!H:H, Run!D:D, "Traditional")'
                ws_report["C5"] = '=AVERAGEIFS(Run!H:H, Run!D:D, "Traditional")' # Coverage Workflow (Same as above)
                ws_report["C6"] = '=AVERAGEIFS(Run!K:K, Run!D:D, "Traditional")'
                ws_report["C7"] = '=AVERAGEIFS(Run!E:E, Run!D:D, "Traditional")'
                ws_report["C8"] = '=SUMIFS(Run!F:F, Run!D:D, "Traditional")'
                ws_report["C9"] = '=SUMIFS(Run!G:G, Run!D:D, "Traditional")'
                ws_report["C10"] = '=COUNTIFS(Run!D:D, "Traditional", Run!J:J, "Đúng")'
                ws_report["C11"] = '=AVERAGEIFS(Run!L:L, Run!D:D, "Traditional")'
                
                # CFG (Column D)
                ws_report["D4"] = '=AVERAGEIFS(Run!H:H, Run!D:D, "Graph-based (CFG)")'
                ws_report["D5"] = '=AVERAGEIFS(Run!H:H, Run!D:D, "Graph-based (CFG)")'
                ws_report["D6"] = '=AVERAGEIFS(Run!K:K, Run!D:D, "Graph-based (CFG)")'
                ws_report["D7"] = '=AVERAGEIFS(Run!E:E, Run!D:D, "Graph-based (CFG)")'
                ws_report["D8"] = '=SUMIFS(Run!F:F, Run!D:D, "Graph-based (CFG)")'
                ws_report["D9"] = '=SUMIFS(Run!G:G, Run!D:D, "Graph-based (CFG)")'
                ws_report["D10"] = '=COUNTIFS(Run!D:D, "Graph-based (CFG)", Run!J:J, "Đúng")'
                ws_report["D11"] = '=AVERAGEIFS(Run!L:L, Run!D:D, "Graph-based (CFG)")'

            # --- SUMMARY / RUN SHEET ---
            if "Run" not in wb.sheetnames:
                ws_run = wb.create_sheet("Run")
            else:
                ws_run = wb["Run"]
                
            # Always force correct headers on row 3
            run_headers = [
                "STT Câu", "RUN_ID", "Tên Business", "Phương pháp", 
                "Thời gian gen", "Input Tokens", "Output Tokens", 
                "Coverage (%)", "Active Node Count", "Accuracy", 
                "Accuracy Rate", "Cyclomatic"
            ]
            for col_idx, header in enumerate(run_headers, 1):
                ws_run.cell(row=3, column=col_idx, value=header)
            
            # Clear old extra columns if any
            for col_idx in range(len(run_headers) + 1, ws_run.max_column + 1):
                ws_run.cell(row=3, column=col_idx, value="")
                
            if clear_old_data and ws_run.max_row >= 4:
                ws_run.delete_rows(4, ws_run.max_row - 3)
                
            def write_run_sheet(mode, results_dict, details_list):
                run_row = -1
                for r in range(4, max(5, ws_run.max_row + 5)):
                    if not ws_run.cell(row=r, column=1).value:
                        run_row = r
                        break
                if run_row == -1: run_row = ws_run.max_row + 1
                
                stt = 1
                for q in details_list:
                    ws_run.cell(row=run_row, column=1, value=stt)
                    ws_run.cell(row=run_row, column=2, value=f"{results['runId']}_{'TRAD' if 'Trad' in mode else 'CFG'}")
                    ws_run.cell(row=run_row, column=3, value=results['businessName'])
                    ws_run.cell(row=run_row, column=4, value=mode)
                    ws_run.cell(row=run_row, column=5, value=results_dict.get(f"{'trad' if 'Trad' in mode else 'cfg'}Time", 0))
                    ws_run.cell(row=run_row, column=6, value=results_dict.get(f"{'trad' if 'Trad' in mode else 'cfg'}InputTokens", 0))
                    ws_run.cell(row=run_row, column=7, value=results_dict.get(f"{'trad' if 'Trad' in mode else 'cfg'}OutputTokens", 0))
                    ws_run.cell(row=run_row, column=8, value=q.get("coverage", 0))
                    ws_run.cell(row=run_row, column=9, value=q.get("activeNodes", 0))
                    ws_run.cell(row=run_row, column=10, value="Đúng" if q.get("isAccurate", False) else "Sai")
                    ws_run.cell(row=run_row, column=11, value=results_dict.get(f"{'trad' if 'Trad' in mode else 'cfg'}Accuracy", 0))
                    ws_run.cell(row=run_row, column=12, value=q.get("cyclomatic", 0))
                    
                    run_row += 1
                    stt += 1

            if method in ["Both", "Traditional"]:
                write_run_sheet("Traditional", results, results.get("tradDetails", []))
            if method in ["Both", "CFG"]:
                write_run_sheet("Graph-based (CFG)", results, results.get("cfgDetails", []))

            # --- DETAILS SHEET ---
            if "Chi tiết Câu hỏi" not in wb.sheetnames:
                ws_details = wb.create_sheet("Chi tiết Câu hỏi")
            else:
                ws_details = wb["Chi tiết Câu hỏi"]
                
            # Always force correct headers on row 3
            details_headers = ["RUN_ID", "Phương Pháp", "STT Câu", "Nội dung câu hỏi"]
            for col_idx, header in enumerate(details_headers, 1):
                ws_details.cell(row=3, column=col_idx, value=header)
                
            # Clear old extra columns if any (like Gateways, etc)
            for col_idx in range(len(details_headers) + 1, ws_details.max_column + 1):
                ws_details.cell(row=3, column=col_idx, value="")
                # Clear entire column data to be safe
                for r in range(4, ws_details.max_row + 1):
                    ws_details.cell(row=r, column=col_idx, value="")

            if clear_old_data and ws_details.max_row >= 4:
                ws_details.delete_rows(4, ws_details.max_row - 3)
                
            def write_details(mode, details_list):
                d_row = -1
                for r in range(4, max(5, ws_details.max_row + 5)):
                    if not ws_details.cell(row=r, column=1).value:
                        d_row = r
                        break
                if d_row == -1: d_row = ws_details.max_row + 1
                
                stt = 1
                for q in details_list:
                    ws_details.cell(row=d_row, column=1, value=f"{results['runId']}_{'TRAD' if 'Trad' in mode else 'CFG'}")
                    ws_details.cell(row=d_row, column=2, value=mode)
                    ws_details.cell(row=d_row, column=3, value=stt)
                    ws_details.cell(row=d_row, column=4, value=q.get("question", ""))
                    d_row += 1
                    stt += 1

            if method in ["Both", "Traditional"]:
                write_details("Traditional", results.get("tradDetails", []))
            if method in ["Both", "CFG"]:
                write_details("Graph-based (CFG)", results.get("cfgDetails", []))

            wb.save(excel_path)
            self.log(f"Đã lưu thành công vào file Excel!")
        except Exception as e:
            self.log(f"[LỖI EXCEL] Không thể ghi file: {str(e)}")
            
    def open_logs_viewer(self):
        log_dir = "benchmark_logs"
        if not os.path.exists(log_dir):
            messagebox.showinfo("Thông báo", "Chưa có dữ liệu log nào được ghi nhận.")
            return
            
        json_files = [f for f in os.listdir(log_dir) if f.endswith(".json")]
        if not json_files:
            messagebox.showinfo("Thông báo", "Chưa có file log JSON nào trong thư mục.")
            return
            
        json_files.sort(reverse=True) # Newest first
        
        viewer = ctk.CTkToplevel(self)
        viewer.title("Trình xem chi tiết Câu Hỏi (Log Viewer)")
        viewer.geometry("1100x700")
        viewer.transient(self)
        viewer.configure(fg_color="#F0F4F8")
        
        # --- TOP FRAME: Select file ---
        top_frame = ctk.CTkFrame(viewer, fg_color="#FFFFFF", corner_radius=8)
        top_frame.pack(fill="x", padx=20, pady=15)
        
        ctk.CTkLabel(top_frame, text="Chọn file log:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=10, pady=10)
        combo_logs = ctk.CTkComboBox(top_frame, values=json_files, width=300, state="readonly")
        combo_logs.pack(side="left", padx=10, pady=10)
        combo_logs.set(json_files[0])
        
        # --- MAIN FRAME ---
        main_frame = ctk.CTkFrame(viewer, fg_color="transparent")
        main_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Left Panel (Questions List)
        left_panel = ctk.CTkScrollableFrame(main_frame, width=250, fg_color="#FFFFFF", corner_radius=8, border_width=1, border_color="#E1E5EB")
        left_panel.pack(side="left", fill="y", padx=(0, 10))
        
        # Right Panel (Details)
        right_panel = ctk.CTkFrame(main_frame, fg_color="#FFFFFF", corner_radius=8, border_width=1, border_color="#E1E5EB")
        right_panel.pack(side="right", fill="both", expand=True)
        
        txt_details = ctk.CTkTextbox(right_panel, font=ctk.CTkFont(family="Consolas", size=14), fg_color="transparent", text_color="#333333", wrap="word")
        txt_details.pack(fill="both", expand=True, padx=15, pady=15)
        
        current_data = {}
        
        def show_question_details(method, idx, q_text):
            txt_details.configure(state="normal")
            txt_details.delete("1.0", "end")
            
            if not current_data or method not in current_data or not current_data[method]:
                txt_details.insert("end", "Không có dữ liệu.")
                txt_details.configure(state="disabled")
                return
                
            data = current_data[method]
            
            # Find data across the 4 APIs
            gen_list = data["generate"].get("generatedQuestionDtos", data["generate"].get("GeneratedQuestionDtos", []))
            cov_list = data["coverage"].get("questionResults", [])
            acc_list = data["accuracy"].get("questionResults", [])
            diff_list = data["difficulty"].get("questionResults", [])
            
            # Match
            gen_match = next((q for q in gen_list if q.get("question", q.get("Question", "")) == q_text), {})
            cov_match = next((q for q in cov_list if q.get("question") == q_text), {})
            acc_match = next((q for q in acc_list if q.get("question") == q_text), {})
            diff_match = next((q for q in diff_list if q.get("question") == q_text), {})
            
            # Formatting output
            out = f"=== CHI TIẾT CÂU HỎI {idx} ({'Code Thô' if method == 'traditional' else 'Đồ thị CFG'}) ===\n\n"
            out += f"❓ CÂU HỎI:\n{q_text}\n\n"
            
            out += f"💡 ĐÁP ÁN (Dự kiến):\n"
            out += f"{gen_match.get('answer', gen_match.get('Answer', 'N/A'))}\n\n"
            
            out += "-" * 50 + "\n\n"
            out += f"📊 1. ĐÁNH GIÁ ĐỘ BAO PHỦ (COVERAGE)\n"
            out += f"   - Coverage Score: {cov_match.get('coverage', 0) * 100:.2f}%\n"
            out += f"   - Active Nodes Count: {cov_match.get('activeNodeCount', 0)}\n"
            out += f"   - Chi tiết Nodes được kích hoạt:\n"
            for node in cov_match.get('activeNodes', []):
                out += f"      + [{node.get('nodeId')}] {node.get('nodeName')}\n"
            
            out += "\n" + "-" * 50 + "\n\n"
            acc_res = acc_match.get("accuracyResult", {})
            out += f"🎯 2. ĐÁNH GIÁ TÍNH ĐÚNG ĐẮN (ACCURACY)\n"
            out += f"   - Kết quả: {'✅ ĐÚNG LOGIC' if acc_res.get('isAccurate') else '❌ SAI LOGIC'}\n"
            out += f"   - Lập luận của AI (Final Verdict):\n"
            verdict = acc_res.get('finalVerdict', 'N/A')
            import textwrap
            for line in textwrap.wrap(verdict, width=80):
                out += f"      {line}\n"
                
            out += "\n" + "-" * 50 + "\n\n"
            diff_res = diff_match.get("difficultyResult", {})
            out += f"🔥 3. ĐÁNH GIÁ ĐỘ PHỨC TẠP (DIFFICULTY)\n"
            out += f"   - Điểm Cyclomatic (Số cạnh Active): {diff_res.get('cyclomaticComplexity', 0)}\n"
            out += f"   - Mức độ đánh giá: {diff_res.get('difficultyLevel', 'N/A')}\n"
            
            txt_details.insert("end", out)
            txt_details.configure(state="disabled")

        def load_selected_log(*args):
            selected = combo_logs.get()
            file_path = os.path.join(log_dir, selected)
            
            # Clear old buttons
            for widget in left_panel.winfo_children():
                widget.destroy()
                
            txt_details.configure(state="normal")
            txt_details.delete("1.0", "end")
            txt_details.configure(state="disabled")
            
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    nonlocal current_data
                    current_data = json.load(f)
                    
                # Add buttons for Traditional
                if current_data.get("traditional"):
                    lbl_t = ctk.CTkLabel(left_panel, text="Code Thô (Traditional)", font=ctk.CTkFont(weight="bold"), text_color="#D83B01")
                    lbl_t.pack(pady=(10, 5), anchor="w", padx=10)
                    
                    t_gen = current_data["traditional"]["generate"].get("generatedQuestionDtos", current_data["traditional"]["generate"].get("GeneratedQuestionDtos", []))
                    for i, q in enumerate(t_gen, 1):
                        q_text = q.get("question", q.get("Question", ""))
                        # Using default arg in lambda to capture loop variables correctly
                        btn = ctk.CTkButton(left_panel, text=f"Câu {i}", fg_color="transparent", text_color="#333333", hover_color="#E1E5EB", anchor="w",
                                            command=lambda m="traditional", idx=i, qt=q_text: show_question_details(m, idx, qt))
                        btn.pack(fill="x", padx=5, pady=2)
                        
                # Add buttons for CFG
                if current_data.get("cfg"):
                    lbl_c = ctk.CTkLabel(left_panel, text="Đồ thị (CFG)", font=ctk.CTkFont(weight="bold"), text_color="#107C41")
                    lbl_c.pack(pady=(15, 5), anchor="w", padx=10)
                    
                    c_gen = current_data["cfg"]["generate"].get("generatedQuestionDtos", current_data["cfg"]["generate"].get("GeneratedQuestionDtos", []))
                    for i, q in enumerate(c_gen, 1):
                        q_text = q.get("question", q.get("Question", ""))
                        btn = ctk.CTkButton(left_panel, text=f"Câu {i}", fg_color="transparent", text_color="#333333", hover_color="#E1E5EB", anchor="w",
                                            command=lambda m="cfg", idx=i, qt=q_text: show_question_details(m, idx, qt))
                        btn.pack(fill="x", padx=5, pady=2)
                        
            except Exception as e:
                txt_details.configure(state="normal")
                txt_details.insert("end", f"Lỗi đọc file: {str(e)}")
                txt_details.configure(state="disabled")
                
        btn_load = ctk.CTkButton(top_frame, text="Mở File & Phân Tích", command=load_selected_log, fg_color="#0066CC")
        btn_load.pack(side="left", padx=10, pady=10)
        
        load_selected_log()

if __name__ == "__main__":
    app = BenchmarkApp()
    app.mainloop()
